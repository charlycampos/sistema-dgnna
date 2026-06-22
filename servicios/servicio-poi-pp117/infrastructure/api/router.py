"""
Router: Módulo POI - PP117
Carga mensual del archivo DATA-POI y generación de reportes DGNNA y PP0117.
Todas las 277 columnas del Excel se almacenan como JSON en la columna 'datos'.
"""

import io
import json
import math
import os
from collections import defaultdict
from datetime import datetime
from typing import List, Optional

import jwt as pyjwt
import pandas as pd
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query, Request
from fastapi.responses import StreamingResponse
from fastapi.security import OAuth2PasswordBearer
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from infrastructure.db.database import get_db
from infrastructure.db.models import PoiCarga, PoiDato

router = APIRouter(prefix="/api/poi-pp117", tags=["poi-pp117"])

SECRET_KEY    = os.getenv("SESSION_SECRET", "dgnna-sistema-dgnna-secret-2026")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="http://localhost:8001/api/auth/login", auto_error=False)

MESES_NOMBRES = ["ENE","FEB","MAR","ABR","MAY","JUN","JUL","AGO","SET","OCT","NOV","DIC"]

CC_DGNNA = {
    "DIRECCIÓN DE ADOPCIONES",
    "DIRECCIÓN DE POLÍTICAS DE NIÑAS, NIÑOS Y ADOLESCENTES",
    "DIRECCIÓN DE PROTECCIÓN ESPECIAL",
    "DIRECCIÓN DE SISTEMAS LOCALES Y DEFENSORÍAS",
    "DIRECCIÓN GENERAL DE NIÑAS, NIÑOS Y ADOLESCENTES",
}
CC_INABIF = {
    "UNIDAD DE ADMINISTRACIÓN",
    "UNIDAD DE DESARROLLO INTEGRAL DE LAS FAMILIAS",
    "UNIDAD DE FORTALECIMIENTO DE SERVICIOS Y COORDINACION TERRITORIAL",
    "UNIDAD DE SERVICIOS DE PROTECCIÓN DE NIÑOS, NIÑAS Y ADOLESCENTES",
    "UNIDAD DE SERVICIOS TERRITORIALES",
}
CC_CONADIS = {"CONADIS"}
CC_PP117   = CC_DGNNA | CC_INABIF | CC_CONADIS

CC_ALIAS = {
    "DIRECCIÓN DE ADOPCIONES": "DA",
    "DIRECCIÓN DE POLÍTICAS DE NIÑAS, NIÑOS Y ADOLESCENTES": "DPNNA",
    "DIRECCIÓN DE PROTECCIÓN ESPECIAL": "DPE",
    "DIRECCIÓN DE SISTEMAS LOCALES Y DEFENSORÍAS": "DSLD",
    "DIRECCIÓN GENERAL DE NIÑAS, NIÑOS Y ADOLESCENTES": "DGNNA",
    "UNIDAD DE ADMINISTRACIÓN": "UA",
    "UNIDAD DE DESARROLLO INTEGRAL DE LAS FAMILIAS": "UDIF",
    "UNIDAD DE FORTALECIMIENTO DE SERVICIOS Y COORDINACION TERRITORIAL": "UFSCT",
    "UNIDAD DE SERVICIOS DE PROTECCIÓN DE NIÑOS, NIÑAS Y ADOLESCENTES": "USPNNA",
    "UNIDAD DE SERVICIOS TERRITORIALES": "UST",
    "CONADIS": "CONADIS",
}

AO_TIPOS = [
    "ATENCIÓN PRELIMINAR DE NIÑAS, NIÑOS Y ADOLESCENTES INGRESADOS AL SERVICIO",
    "ELABORACIÓN Y APROBACIÓN DEL PLAN DE TRABAJO INDIVIDUAL DE NIÑAS, NIÑOS Y ADOLESCENTES CON ACOGIMIENTO FAMILIAR",
    "ELABORACIÓN Y APROBACIÓN DEL PLAN DE TRABAJO INDIVIDUAL DE NIÑAS, NIÑOS Y ADOLESCENTES CON ACOGIMIENTO RESIDENCIAL",
    "ELABORACIÓN Y APROBACIÓN DEL PLAN DE TRABAJO INDIVIDUAL DE NIÑAS, NIÑOS Y ADOLESCENTES EN RIESGO DE DESPROTECCION FAMILIAR",
    "EVALUACIÓN SOCIOFAMILIAR DE NIÑAS, NIÑOS Y ADOLESCENTES INGRESADOS AL SERVICIO",
]
AO_TYPOS = {
    "ATENCIÓN PRELIMINAR DE NIÑAS, NIÑOS Y ADOELSCENTES INGRESADOS AL SERVICIO":
        "ATENCIÓN PRELIMINAR DE NIÑAS, NIÑOS Y ADOLESCENTES INGRESADOS AL SERVICIO",
    "ATENCIÓN PRELIMINAR DE NIÑAS, NIÑOS Y ADOELSCENTES INGRESADOS AL SERVICIO":
        "ATENCIÓN PRELIMINAR DE NIÑAS, NIÑOS Y ADOLESCENTES INGRESADOS AL SERVICIO",
}

META_PPTAL_UPE_AO1 = {
    "UPE AMAZONAS": 39, "UPE ANCASH": 40, "UPE APURIMAC": 41, "UPE AREQUIPA": 42,
    "UPE AYACUCHO": 43, "UPE CAJAMARCA": 44, "UPE CUSCO": 46, "UPE HUANCAVELICA": 47,
    "UPE HUANUCO": 48, "UPE ICA": 49, "UPE JUNIN": 50, "UPE LA LIBERTAD": 51,
    "UPE LAMBAYEQUE": 52, "UPE LIMA": 53, "UPE LIMA ESTE": 56, "UPE LIMA NORTE CALLAO": 45,
    "UPE LIMA SUR": 55, "UPE LORETO": 58, "UPE MADRE DE DIOS": 59, "UPE MOQUEGUA": 60,
    "UPE PIURA": 61, "UPE PUNO": 62, "UPE SAN MARTIN": 63, "UPE TACNA": 64,
    "UPE TUMBES": 65, "UPE UCAYALI": 66
}


# ─── Auth helpers ────────────────────────────────────────────────

def _get_usuario(token: str = Depends(oauth2_scheme)) -> dict:
    if not token:
        raise HTTPException(status_code=401, detail="Sesión inválida o expirada")
    try:
        payload = pyjwt.decode(token, SECRET_KEY, algorithms=["HS256"])
        return payload
    except pyjwt.PyJWTError:
        raise HTTPException(status_code=401, detail="Sesión inválida o expirada")


def _nombre_usuario(token: str = Depends(oauth2_scheme)) -> str:
    if not token:
        return ""
    try:
        payload = pyjwt.decode(token, SECRET_KEY, algorithms=["HS256"])
        return payload.get("nombre", payload.get("sub", ""))
    except Exception:
        return ""


# ─── Helpers generales ───────────────────────────────────────────

def _pct(a, b):
    if b and b != 0:
        return round(a / b * 100, 2)
    return None

def _get_carga_o_404(db, mes, anio):
    c = db.query(PoiCarga).filter(PoiCarga.mes == mes, PoiCarga.anio == anio).first()
    if not c:
        raise HTTPException(404, f"No hay datos cargados para {MESES_NOMBRES[mes-1]} {anio}")
    return c

def _safe_json(val):
    """Convierte un valor de pandas a tipo JSON-serializable."""
    if val is None:
        return None
    try:
        if isinstance(val, float) and math.isnan(val):
            return None
        if isinstance(val, (int, float)):
            return float(val)
        s = str(val).strip()
        return s if s and s.lower() not in ("nan", "none", "") else None
    except Exception:
        return None

def _parse_datos(f: PoiDato) -> dict:
    """Lee el JSON de datos almacenado en la fila."""
    if f.datos:
        try:
            return json.loads(f.datos)
        except Exception:
            return {}
    return {}

def _jval(d: dict, key: str, default: float = 0.0) -> float:
    """Obtiene un valor numérico del JSON."""
    v = d.get(key)
    try:
        return float(v) if v is not None else default
    except (TypeError, ValueError):
        return default

def _jstr(d: dict, key: str) -> Optional[str]:
    """Obtiene un valor string del JSON."""
    v = d.get(key)
    if v is None:
        return None
    s = str(v).strip()
    return s if s and s.lower() not in ("nan", "none") else None


# ─── GET /periodos  (alias: /meses) ──────────────────────────────

def _listar_periodos(db):
    cargas = db.query(PoiCarga).order_by(PoiCarga.anio.desc(), PoiCarga.mes.desc()).all()
    return [{"id": c.id, "mes": c.mes, "anio": c.anio,
             "mesNombre": MESES_NOMBRES[c.mes-1],
             "nombreArchivo": c.nombreArchivo,
             "totalFilas": c.totalFilas,
             "cargadoPor": c.cargadoPor,
             "createdAt": c.createdAt.isoformat() if c.createdAt else None}
            for c in cargas]

@router.get("/periodos")
def listar_periodos(db: Session = Depends(get_db), _: dict = Depends(_get_usuario)):
    return _listar_periodos(db)

@router.get("/meses")
def listar_meses(db: Session = Depends(get_db), _: dict = Depends(_get_usuario)):
    return _listar_periodos(db)


# ─── POST /cargar ────────────────────────────────────────────────

@router.post("/cargar", status_code=201)
async def cargar_archivo(
    mes: int = Query(..., ge=1, le=12),
    anio: int = Query(..., ge=2020, le=2099),
    reemplazar: bool = Query(False),
    archivo: UploadFile = File(...),
    db: Session = Depends(get_db),
    nombre: str = Depends(_nombre_usuario),
):
    existente = db.query(PoiCarga).filter(PoiCarga.mes == mes, PoiCarga.anio == anio).first()
    if existente and not reemplazar:
        raise HTTPException(409, f"Ya existe carga para {MESES_NOMBRES[mes-1]} {anio}. Use reemplazar=true.")

    contenido = await archivo.read()
    try:
        df = pd.read_excel(io.BytesIO(contenido), sheet_name="POI_Por_ActividadOperativaAnual")
    except Exception as e:
        raise HTTPException(400, f"Error leyendo Excel: {e}")

    if existente:
        db.delete(existente); db.commit()

    carga = PoiCarga(mes=mes, anio=anio, nombreArchivo=archivo.filename,
                     totalFilas=len(df), cargadoPor=nombre)
    db.add(carga); db.flush()

    registros = []
    columnas = list(df.columns)

    for _, row in df.iterrows():
        cc = str(row.get("Centro de Costo") or "").strip()
        if cc not in CC_PP117:
            continue

        # Serializar TODAS las columnas como JSON
        row_dict = {col: _safe_json(row.get(col)) for col in columnas}

        dato = PoiDato(
            cargaId      = carga.id,
            centroCosto  = cc,
            actividadOp  = _safe_json(row.get("Actividad Operativa")),
            departamento = _safe_json(row.get("Departamento Nombre UBIGEO")),
            codAO        = _safe_json(row.get("Actividad Operativa ID")),
            datos        = json.dumps(row_dict, ensure_ascii=False),
        )
        registros.append(dato)

    db.bulk_save_objects(registros); db.commit()
    return {"mensaje": f"Carga exitosa: {MESES_NOMBRES[mes-1]} {anio}",
            "filasGuardadas": len(registros), "totalFilas": len(df)}


# ─── Agregación desde JSON ───────────────────────────────────────

def _agregar_filas(filas: List[PoiDato], mes_eval: int) -> List[dict]:
    grupos: dict = defaultdict(lambda: {
        "centroCosto":"","categoriaId":"","categoria":"",
        "actPresupId":"","actPresup":"",
        "codAO":"","actividadOp":"","unidadMedida":"","motivoMes":"",
        "fRe":[0.0]*12,"fReTotal":0.0,
        "fSe":[0.0]*12,"fSeTotal":0.0,
        "fnReTotal":0.0,"fnSeTotal":0.0,
    })
    for f in filas:
        d = _parse_datos(f)
        key = (f.centroCosto, f.codAO or _jstr(d, "Actividad Operativa ID") or "")
        g = grupos[key]
        g["centroCosto"]  = f.centroCosto or ""
        g["categoriaId"]  = _jstr(d, "Categoria ID") or ""
        g["categoria"]    = _jstr(d, "Categoria") or g["categoria"]
        g["actPresupId"]  = _jstr(d, "Actividad Presupuestal ID") or ""
        g["actPresup"]    = _jstr(d, "Actividad Presupuestal") or g["actPresup"]
        g["codAO"]        = f.codAO or ""
        g["actividadOp"]  = f.actividadOp or ""
        g["unidadMedida"] = _jstr(d, "Unidad de Medida") or g["unidadMedida"]
        motivo_key = f"Motivo(SE) {mes_eval:02d}"
        g["motivoMes"]    = _jstr(d, motivo_key) or g["motivoMes"]
        for i in range(12):
            g["fRe"][i] += _jval(d, f"F(RE) {i+1:02d}")
            g["fSe"][i] += _jval(d, f"F(SE) {i+1:02d}")
        g["fReTotal"]  += _jval(d, "F(RE) Total")
        g["fSeTotal"]  += _jval(d, "F(SE) Total")
        g["fnReTotal"] += _jval(d, "Fn(RE) Total")
        g["fnSeTotal"] += _jval(d, "Fn(SE) Total")

    resultado = []
    for g in grupos.values():
        # Determinar si los datos de la fila son acumulados o incrementales
        # Comparamos el último mes F(RE) 12 con F(RE) Total
        is_running_sum = abs(g["fRe"][11] - g["fReTotal"]) <= 1.0

        if is_running_sum:
            prog_acum = g["fRe"][mes_eval - 1]
            ejec_acum = g["fSe"][mes_eval - 1]
        else:
            prog_acum = sum(g["fRe"][:mes_eval])
            ejec_acum = sum(g["fSe"][:mes_eval])
            
        prog_mes1 = g["fRe"][0]
        ejec_mes1 = g["fSe"][0]
        resultado.append({
            "centroCosto":   g["centroCosto"],
            "ccAlias":       CC_ALIAS.get(g["centroCosto"], g["centroCosto"][:6]),
            "categoriaId":   g["categoriaId"],
            "categoria":     g["categoria"],
            "actPresupId":   g["actPresupId"],
            "actPresup":     g["actPresup"],
            "codAO":         g["codAO"],
            "actividadOp":   g["actividadOp"],
            "unidadMedida":  g["unidadMedida"],
            "progAnual":     round(g["fReTotal"], 2),
            "progFinAnual":  round(g["fnReTotal"], 2),
            "progMeses":     [round(v, 2) for v in g["fRe"]],
            "ejecMeses":     [round(v, 2) for v in g["fSe"]],
            "progMes1":      round(prog_mes1, 2),
            "ejecMes1":      round(ejec_mes1, 2),
            "pctMes1":       _pct(ejec_mes1, prog_mes1),
            "progAcum":      round(prog_acum, 2),
            "ejecAcum":      round(ejec_acum, 2),
            "pctAvance":     _pct(ejec_acum, prog_acum),
            "pim":           round(g["fnReTotal"], 2),
            "devengado":     round(g["fnSeTotal"], 2),
            "pctFinanciero": _pct(g["fnSeTotal"], g["fnReTotal"]),
            "motivoMes":     g["motivoMes"],
        })
    return resultado


# ─── GET /dashboard ──────────────────────────────────────────────

@router.get("/dashboard")
def dashboard(
    mes: int  = Query(..., ge=1, le=12),
    anio: int = Query(..., ge=2020),
    tipo: str = Query("dgnna"),
    db: Session = Depends(get_db),
    _: dict = Depends(_get_usuario),
):
    carga = _get_carga_o_404(db, mes, anio)
    cc_filtro = CC_DGNNA if tipo == "dgnna" else CC_PP117
    filas = db.query(PoiDato).filter(PoiDato.cargaId == carga.id,
                                      PoiDato.centroCosto.in_(cc_filtro)).all()
    datos = _agregar_filas(filas, mes)

    prog_total = sum(d["progAcum"]  for d in datos)
    ejec_total = sum(d["ejecAcum"]  for d in datos)
    pim_total  = sum(d["pim"]       for d in datos)
    dev_total  = sum(d["devengado"] for d in datos)

    verde    = sum(1 for d in datos if (d["pctAvance"] or 0) >= 95)
    amarillo = sum(1 for d in datos if 75 <= (d["pctAvance"] or 0) < 95)
    rojo     = sum(1 for d in datos if d["pctAvance"] is not None and d["pctAvance"] < 75)

    cc_map: dict = defaultdict(lambda: {"alias":"","prog":0,"ejec":0,"pim":0,"dev":0})
    for d in datos:
        cc = d["centroCosto"]
        cc_map[cc]["alias"] = d["ccAlias"]
        cc_map[cc]["prog"] += d["progAcum"]
        cc_map[cc]["ejec"] += d["ejecAcum"]
        cc_map[cc]["pim"]  += d["pim"]
        cc_map[cc]["dev"]  += d["devengado"]

    por_cc = [
        {"cc": CC_ALIAS.get(cc, cc[:8]), "ccCompleto": cc,
         "pctFisico":     round(v["ejec"]/v["prog"]*100, 1) if v["prog"] else 0,
         "pctFinanciero": round(v["dev"]/v["pim"]*100, 1)   if v["pim"]  else 0,
         "ejecAcum": round(v["ejec"], 2), "progAcum": round(v["prog"], 2)}
        for cc, v in cc_map.items()
    ]
    por_cc.sort(key=lambda x: x["pctFisico"], reverse=True)

    mensual = [
        {"mes": MESES_NOMBRES[i],
         "prog": round(sum(d["progMeses"][i] for d in datos), 2),
         "ejec": round(sum(d["ejecMeses"][i] for d in datos), 2)}
        for i in range(mes)
    ]

    return {
        "mes": mes, "anio": anio, "mesNombre": MESES_NOMBRES[mes-1],
        "kpis": {
            "totalAOs": len(datos),
            "pctAvanceFisico": _pct(ejec_total, prog_total) or 0,
            "pimTotal": round(pim_total, 2),
            "devengadoTotal": round(dev_total, 2),
            "pctFinanciero": _pct(dev_total, pim_total) or 0,
        },
        "semaforo": {"verde": verde, "amarillo": amarillo, "rojo": rojo},
        "porCC": por_cc,
        "mensual": mensual,
    }


# ─── TABLA 01 - Lógica DGNNA ───────────────────────────────────────

TABLA01_AOS_PERMITIDAS = {
    "DA": [
        "Integración de niñas, niños y adolescentes a un entorno familiar definitivo.",
        "Seguimiento a niñas, niños y adolescentes integrados a una familia adoptiva.",
        "Evaluación a familias para la adopción de NNA en desprotección familiar"
    ],
    "DSLD": [
        "Intervención lúdica y espacios seguros para el fortalecimiento de capacidades de niñas, niños y adolescentes - Multisectorial",
        "Intervención lúdica y espacios seguros para el fortalecimiento de capacidades de niñas, niños y adolescentes",
        "Acreditación y supervisión de DEMUNA",
        "Fortalecimiento de las defensorías municipales de la niña, niño y adolescente - DEMUNA",
        "Fortalecimiento de los Consejos Consultivos y participativos de Niñas, Niños y Adolescentes - CCONNA",
        "Implementación de la estrategia Ponte en #ModoNiñez"
    ],
    "DPNNA": [
        "Supervisión y/o acreditación de centros de acogida residencial públicos y privados",
        "Implementación seguimiento y evaluación de la Política Nacional Multisectorial para las Niñas, Niños y Adolescentes",
        "Fortalecimiento de los servicios de atención a niñas, niños y adolescentes vulnerables al delito de explotación sexual en el ámbito nacional.",
        "Seguimiento a la atención y reintegración a niñas, niños y adolescentes afectados por el delito de trata de personas en el ámbito nacional."
    ],
    "DGNNA": [
        "Gestión del Programa",
        "Fortalecimiento del sistema de protección de niñas, niños y adolescentes",
        "Atención de restitución del ejercicio de derechos de niñas, niños y adolescentes",
        "Implementación de la Estrategia Multisectorial para la prevención de la violencia sexual que afecta a niñas, niños y adolescentes PREVENIR PARA PROTEGER"
    ],
    "UPE": [
        "Atención preliminar de niñas, niños y adolescentes ingresados al servicio",
        "Elaboración y aprobación del plan de trabajo individual de niñas, niños y adolescentes en riesgo de desproteccion familiar",
        "Evaluación sociofamiliar de niñas, niños y adolescentes ingresados al servicio",
        "Elaboración y aprobación del plan de trabajo individual de niñas, niños y adolescentes con acogimiento residencial",
        "Elaboración y aprobación del plan de trabajo individual de niñas, niños y adolescentes con acogimiento familiar"
    ],
    "DPE": [
        "Gestión del Programa",
        "Atención de llamadas a través de la línea ANNA 1810",
        "Supervisión a las UPE",
        "Evaluación, declaración de capacidad y capacitación de la familia para el acogimiento familiar",
        "Subvención económica por acogimiento familiar a las niñas, niños y adolescentes"
    ]
}

def _agregar_filas_tabla01(filas: List[PoiDato], mes_eval: int) -> List[dict]:
    import unicodedata
    def normalizar(s):
        if not s: return ""
        s = str(s).replace("–", "-")
        return unicodedata.normalize('NFKD', s).encode('ASCII', 'ignore').decode('utf-8').upper().strip()

    DB_CC_A_GRUPO = {
        "DIRECCIÓN DE ADOPCIONES": "DA",
        "DIRECCIÓN DE SISTEMAS LOCALES Y DEFENSORÍAS": "DSLD",
        "DIRECCIÓN DE POLÍTICAS DE NIÑAS, NIÑOS Y ADOLESCENTES": "DPNNA",
        "DIRECCIÓN GENERAL DE NIÑAS, NIÑOS Y ADOLESCENTES": "DGNNA",
    }
    upe_aos_norm = {normalizar(a): a for a in TABLA01_AOS_PERMITIDAS["UPE"]}
    dpe_aos_norm = {normalizar(a): a for a in TABLA01_AOS_PERMITIDAS["DPE"]}
    cc_permitidas_norm = {cc: {normalizar(a): a for a in acts} for cc, acts in TABLA01_AOS_PERMITIDAS.items()}

    grupos = {}
    for f in filas:
        d = _parse_datos(f)
        cc_db = f.centroCosto
        
        ao_nombre = (f.actividadOp or "").strip().upper()
        ao_nombre = ao_nombre.replace("ADOELSCENTES", "ADOLESCENTES")
        ao_norm = normalizar(ao_nombre)
        
        cc_group = None
        ao_oficial = None
        
        if cc_db == "DIRECCIÓN DE PROTECCIÓN ESPECIAL":
            for k_norm, k_oficial in upe_aos_norm.items():
                if ao_norm.startswith(k_norm):
                    cc_group = "UPE"
                    ao_oficial = k_oficial
                    break
            if not cc_group:
                for k_norm, k_oficial in dpe_aos_norm.items():
                    if ao_norm.startswith(k_norm):
                        cc_group = "DPE"
                        ao_oficial = k_oficial
                        break
            if not cc_group:
                continue
        else:
            cc_group = DB_CC_A_GRUPO.get(cc_db)
            if not cc_group: continue
            for k_norm, k_oficial in cc_permitidas_norm[cc_group].items():
                if ao_norm.startswith(k_norm):
                    ao_oficial = k_oficial
                    break
            if not ao_oficial:
                continue
        
        k = (cc_group, ao_oficial)
        if k not in grupos:
            grupos[k] = {
                "centroCosto": cc_group,
                "codAO": "",
                "actividadOp": ao_oficial,
                "unidadMedida": _jstr(d, "Unidad de Medida") or "",
                "motivoMes": "",
                "fRe": [0.0]*12, "fReTotal": 0.0,
                "fSe": [0.0]*12, "fSeTotal": 0.0,
                "fnReTotal": 0.0, "fnSeTotal": 0.0,
            }
        
        g = grupos[k]
        g["unidadMedida"] = g["unidadMedida"] or _jstr(d, "Unidad de Medida") or ""
        motivo_key = f"Motivo(SE) {mes_eval:02d}"
        g["motivoMes"]    = _jstr(d, motivo_key) or g["motivoMes"]
        for i in range(12):
            g["fRe"][i] += _jval(d, f"F(RE) {i+1:02d}")
            g["fSe"][i] += _jval(d, f"F(SE) {i+1:02d}")
        g["fReTotal"]  += _jval(d, "F(RE) Total")
        g["fSeTotal"]  += _jval(d, "F(SE) Total")
        g["fnReTotal"] += _jval(d, "Fn(RE) Total")
        g["fnSeTotal"] += _jval(d, "Fn(SE) Total")

    resultado = []
    ALIAS_A_NOMBRE = {
        "DA": "DIRECCIÓN DE ADOPCIONES",
        "DSLD": "DIRECCIÓN DE SISTEMAS LOCALES Y DEFENSORÍAS",
        "DPNNA": "DIRECCIÓN DE POLÍTICAS DE NIÑAS, NIÑOS Y ADOLESCENTES",
        "DGNNA": "DIRECCIÓN GENERAL DE NIÑAS, NIÑOS Y ADOLESCENTES",
        "UPE": "UNIDAD DE PROTECCIÓN ESPECIAL (UPE)",
        "DPE": "DIRECCIÓN DE PROTECCIÓN ESPECIAL",
    }
    
    META_PPTAL_MAPPING = {
        ("DPE", "Gestión del Programa"): 3,
        ("DGNNA", "Gestión del Programa"): 4,
        "Fortalecimiento del sistema de protección de niñas, niños y adolescentes": 156,
        "Atención de restitución del ejercicio de derechos de niñas, niños y adolescentes": 158,
        "Implementación de la Estrategia Multisectorial para la prevención de la violencia sexual que afecta a niñas, niños y adolescentes PREVENIR PARA PROTEGER": 196,
        "Atención de llamadas a través de la línea ANNA 1810": 187,
        "Supervisión a las UPE": 67,
        "Evaluación, declaración de capacidad y capacitación de la familia para el acogimiento familiar": 94,
        "Subvención económica por acogimiento familiar a las niñas, niños y adolescentes": 188,
        "Subvención económica para acogimiento familiar a las niñas, niños y adolescentes": 188,
        "Atención preliminar de niñas, niños y adolescentes ingresados al servicio": "39-66",
        "Elaboración y aprobación del plan de trabajo individual de niñas, niños y adolescentes con acogimiento familiar": "69-93",
        "Elaboración y aprobación del plan de trabajo individual de niñas, niños y adolescentes con acogimiento residencial": "0",
        "Elaboración y aprobación del plan de trabajo individual de niñas, niños y adolescentes en riesgo de desproteccion familiar": "14-38",
        "Evaluación sociofamiliar de niñas, niños y adolescentes ingresados al servicio": "0",
        "Integración de niñas, niños y adolescentes a un entorno familiar definitivo.": 96,
        "Seguimiento a niñas, niños y adolescentes integrados a una familia adoptiva.": 97,
        "Evaluación a familias para la adopción de NNA en desprotección familiar": 98,
        "Intervención lúdica y espacios seguros para el fortalecimiento de capacidades de niñas, niños y adolescentes - Multisectorial": 6,
        "Intervención lúdica y espacios seguros para el fortalecimiento de capacidades de niñas, niños y adolescentes": 5,
        "Acreditación y supervisión de DEMUNA": 68,
        "Fortalecimiento de las defensorías municipales de la niña, niño y adolescente - DEMUNA": 155,
        "Fortalecimiento de los Consejos Consultivos y participativos de Niñas, Niños y Adolescentes - CCONNA": 159,
        "Implementación de la estrategia Ponte en #ModoNiñez": 157,
        "Supervisión y/o acreditación de centros de acogida residencial públicos y privados": 95,
        "Implementación seguimiento y evaluación de la Política Nacional Multisectorial para las Niñas, Niños y Adolescentes": 160,
        "Fortalecimiento de los servicios de atención a niñas, niños y adolescentes vulnerables al delito de explotación sexual en el ámbito nacional.": 186,
        "Seguimiento a la atención y reintegración a niñas, niños y adolescentes afectados por el delito de trata de personas en el ámbito nacional.": 180,
    }

    for cc, acts in TABLA01_AOS_PERMITIDAS.items():
        for ao in acts:
            k = (cc, ao)
            if k in grupos:
                g = grupos[k]
                # Determinar si los datos de la fila son acumulados o incrementales
                # Comparamos el último mes F(RE) 12 con F(RE) Total
                # Usamos un umbral del 50% porque puede haber ligeros errores de digitación en el Excel (ej. 25680 vs 25790)
                is_running_sum = (g["fReTotal"] > 0) and (g["fRe"][11] >= g["fReTotal"] * 0.5)

                if is_running_sum:
                    prog_acum = g["fRe"][mes_eval - 1]
                    ejec_acum = g["fSe"][mes_eval - 1]
                else:
                    prog_acum = sum(g["fRe"][:mes_eval])
                    ejec_acum = sum(g["fSe"][:mes_eval])
                    
                prog_mes1 = g["fRe"][0]
                ejec_mes1 = g["fSe"][0]
                resultado.append({
                    "centroCosto":   g["centroCosto"],
                    "ccAlias":       ALIAS_A_NOMBRE.get(g["centroCosto"], g["centroCosto"]),
                    "codAO":         g["codAO"],
                    "actividadOp":   g["actividadOp"],
                    "unidadMedida":  g["unidadMedida"],
                    "progAnual":     round(g["fReTotal"], 2),
                    "progFinAnual":  round(g["fnReTotal"], 2),
                    "progMeses":     [round(v, 2) for v in g["fRe"]],
                    "ejecMeses":     [round(v, 2) for v in g["fSe"]],
                    "progMes1":      round(prog_mes1, 2),
                    "ejecMes1":      round(ejec_mes1, 2),
                    "pctMes1":       _pct(ejec_mes1, prog_mes1),
                    "progAcum":      round(prog_acum, 2),
                    "ejecAcum":      round(ejec_acum, 2),
                    "pctAvance":     _pct(ejec_acum, prog_acum),
                    "pim":           round(g["fnReTotal"], 2),
                    "devengado":     round(g["fnSeTotal"], 2),
                    "pctFinanciero": _pct(g["fnSeTotal"], g["fnReTotal"]),
                    "metaPptal":     META_PPTAL_MAPPING.get((cc, g["actividadOp"])) or META_PPTAL_MAPPING.get(g["actividadOp"]),
                    "motivoMes":     g["motivoMes"],
                })
    return resultado


# ─── GET /dgnna ──────────────────────────────────────────────────

@router.get("/dgnna")
def datos_dgnna(
    mes: int = Query(..., ge=1, le=12),
    anio: int = Query(..., ge=2020),
    db: Session = Depends(get_db),
    _: dict = Depends(_get_usuario),
):
    carga = _get_carga_o_404(db, mes, anio)
    filas = db.query(PoiDato).filter(PoiDato.cargaId == carga.id,
                                      PoiDato.centroCosto.in_(CC_DGNNA)).all()
    datos = _agregar_filas_tabla01(filas, mes)
    cc_grupos: dict = defaultdict(list)
    for d in datos:
        cc_grupos[d["centroCosto"]].append(d)
    
    # Asegurar el orden: DA, DSLD, DPNNA, DGNNA, UPE, DPE
    orden = ["DA", "DSLD", "DPNNA", "DGNNA", "UPE", "DPE"]
    por_cc = []
    for cc_key in orden:
        if cc_key in cc_grupos:
            por_cc.append({
                "cc": cc_key,
                "ccAlias": cc_grupos[cc_key][0]["ccAlias"] if cc_grupos[cc_key] else cc_key,
                "actividades": cc_grupos[cc_key]
            })

    return {
        "mes": mes, "anio": anio, "mesNombre": MESES_NOMBRES[mes-1],
        "datos": datos,
        "porCC": por_cc,
    }



# ─── GET /dgnna-upe ──────────────────────────────────────────────

@router.get("/dgnna-upe")
def datos_dgnna_upe(
    mes: int = Query(..., ge=1, le=12),
    anio: int = Query(..., ge=2020),
    db: Session = Depends(get_db),
    _: dict = Depends(_get_usuario),
):
    carga = _get_carga_o_404(db, mes, anio)
    filas = db.query(PoiDato).filter(
        PoiDato.cargaId == carga.id,
        PoiDato.centroCosto.in_(CC_DGNNA),
    ).all()

    import unicodedata
    def norm_upe(s):
        return ''.join(c for c in unicodedata.normalize('NFD', str(s).upper().strip()) if unicodedata.category(c) != 'Mn')

    def init_upes(ao_name):
        d = {}
        for upe in META_PPTAL_UPE_AO1.keys():
            m_pptal = META_PPTAL_UPE_AO1.get(upe) if ao_name == "ATENCIÓN PRELIMINAR DE NIÑAS, NIÑOS Y ADOLESCENTES INGRESADOS AL SERVICIO" else None
            d[norm_upe(upe)] = {
                "departamento": upe, "progAnual": 0.0, "progAcum": 0.0, "ejecAcum": 0.0,
                "pctAvance": None, "pim": 0.0, "devengado": 0.0, "pctFinanciero": None,
                "metaPptal": m_pptal, "sinDatos": True
            }
        return d

    ao_grupos: dict = {t: {"codAO": t, "actividadOp": t, "unidadMedida": "", "upes_dict": init_upes(t)} for t in AO_TIPOS}

    for f in filas:
        d = _parse_datos(f)
        ao_nombre = (f.actividadOp or "").strip()

        # Separar prefijo y sufijo UPE
        if " - " in ao_nombre:
            ao_prefijo = ao_nombre.split(" - ", 1)[0].strip()
            upe_sufijo = ao_nombre.split(" - ", 1)[1].strip()
        else:
            ao_prefijo = ao_nombre
            upe_sufijo = ""

        # Corregir typos
        ao_prefijo = AO_TYPOS.get(ao_prefijo, ao_prefijo)

        if ao_prefijo not in ao_grupos:
            continue

        g = ao_grupos[ao_prefijo]
        g["unidadMedida"] = g["unidadMedida"] or (_jstr(d, "Unidad de Medida") or "")

        # Nombre UPE
        if upe_sufijo.upper().startswith("UPE "):
            upe_nombre = upe_sufijo.upper()
        elif upe_sufijo:
            upe_nombre = f"UPE {upe_sufijo.upper()}"
        else:
            dep = (f.departamento or "").upper()
            upe_nombre = f"UPE {dep}" if dep else "SIN UPE"

        # Determinar si la fila es acumulada o incremental
        prog_anual = _jval(d, "F(RE) Total")
        fre_12 = _jval(d, "F(RE) 12")
        is_running_sum = (prog_anual > 0) and (fre_12 >= prog_anual * 0.5)

        if is_running_sum:
            prog_acum = _jval(d, f"F(RE) {mes:02d}")
            ejec_acum = _jval(d, f"F(SE) {mes:02d}")
        else:
            prog_acum = sum(_jval(d, f"F(RE) {i:02d}") for i in range(1, mes + 1))
            ejec_acum = sum(_jval(d, f"F(SE) {i:02d}") for i in range(1, mes + 1))
            
        pim        = _jval(d, "Fn(RE) Total")
        devengado  = _jval(d, "Fn(SE) Total")
        sin_datos  = (prog_anual == 0 and ejec_acum == 0)

        import unicodedata
        def norm_upe(s):
            return ''.join(c for c in unicodedata.normalize('NFD', str(s).upper().strip()) if unicodedata.category(c) != 'Mn')

        nupe = norm_upe(upe_nombre)
        if nupe not in g["upes_dict"]:
            m_pptal = META_PPTAL_UPE_AO1.get(upe_nombre) if ao_prefijo == "ATENCIÓN PRELIMINAR DE NIÑAS, NIÑOS Y ADOLESCENTES INGRESADOS AL SERVICIO" else None
            g["upes_dict"][nupe] = {
                "departamento": upe_nombre, "progAnual": 0.0, "progAcum": 0.0, "ejecAcum": 0.0,
                "pctAvance": None, "pim": 0.0, "devengado": 0.0, "pctFinanciero": None,
                "metaPptal": m_pptal, "sinDatos": True
            }
        
        u = g["upes_dict"][nupe]
        u["progAnual"] += round(prog_anual, 2)
        u["progAcum"] += round(prog_acum, 2)
        u["ejecAcum"] += round(ejec_acum, 2)
        u["pim"] += round(pim, 2)
        u["devengado"] += round(devengado, 2)
        u["sinDatos"] = u["sinDatos"] and sin_datos
        u["pctAvance"] = _pct(u["ejecAcum"], u["progAcum"])
        u["pctFinanciero"] = _pct(u["devengado"], u["pim"])

    # Ordenar UPEs alfabéticamente dentro de cada AO
    for g in ao_grupos.values():
        g["upes"] = list(g["upes_dict"].values())
        g["upes"].sort(key=lambda u: u["departamento"])
        del g["upes_dict"]

    return {
        "mes": mes, "anio": anio, "mesNombre": MESES_NOMBRES[mes-1],
        "actividades": [v for v in ao_grupos.values() if v["upes"]],
    }


# ─── GET /estructura-pp117 ───────────────────────────────────────

@router.get("/estructura-pp117")
def estructura_pp117(
    mes: int     = Query(..., ge=1, le=12),
    anio: int    = Query(..., ge=2020),
    entidad: str = Query("sector"),
    db: Session  = Depends(get_db),
    _: dict      = Depends(_get_usuario),
):
    carga = _get_carga_o_404(db, mes, anio)
    cc_map = {"dgnna": CC_DGNNA, "inabif": CC_INABIF, "conadis": CC_CONADIS, "sector": CC_PP117}
    filas = db.query(PoiDato).filter(PoiDato.cargaId == carga.id,
                                      PoiDato.centroCosto.in_(cc_map.get(entidad, CC_PP117))).all()
    datos = _agregar_filas(filas, mes)
    cat_map: dict = defaultdict(lambda: {
        "categoriaId":"","categoria":"","totalProg":0,"totalEjec":0,
        "aps": defaultdict(lambda: {"actPresupId":"","actPresup":"","totalProg":0,"totalEjec":0,"centros":[]})
    })
    for d in datos:
        cat_id = d["categoriaId"] or "S/C"; ap_id = d["actPresupId"] or "S/AP"
        cat = cat_map[cat_id]
        cat["categoriaId"] = cat_id; cat["categoria"] = d["categoria"] or cat_id
        cat["totalProg"] += d["progAcum"]; cat["totalEjec"] += d["ejecAcum"]
        ap = cat["aps"][ap_id]
        ap["actPresupId"] = ap_id; ap["actPresup"] = d["actPresup"] or ap_id
        ap["totalProg"] += d["progAcum"]; ap["totalEjec"] += d["ejecAcum"]
        ap["centros"].append(d)

    categorias = []
    for cat in cat_map.values():
        aps = [{"actPresupId": ap["actPresupId"], "actPresup": ap["actPresup"],
                "totalProg": round(ap["totalProg"],2), "totalEjec": round(ap["totalEjec"],2),
                "pctAvance": _pct(ap["totalEjec"], ap["totalProg"]), "centros": ap["centros"]}
               for ap in cat["aps"].values()]
        categorias.append({
            "categoriaId": cat["categoriaId"], "categoria": cat["categoria"],
            "totalProg": round(cat["totalProg"],2), "totalEjec": round(cat["totalEjec"],2),
            "pctAvance": _pct(cat["totalEjec"], cat["totalProg"]), "aps": aps,
        })
    return {"mes": mes, "anio": anio, "mesNombre": MESES_NOMBRES[mes-1],
            "entidad": entidad, "categorias": categorias}


# ─── DELETE /carga/{id} ──────────────────────────────────────────

@router.delete("/carga/{id}", status_code=204)
def eliminar_carga(id: str, db: Session = Depends(get_db), _: dict = Depends(_get_usuario)):
    c = db.query(PoiCarga).filter(PoiCarga.id == id).first()
    if not c:
        raise HTTPException(404, "Carga no encontrada")
    db.delete(c); db.commit()


# ─── Helpers Excel ───────────────────────────────────────────────

_F_AZUL_OSC   = PatternFill("solid", fgColor="1F4E79")
_F_AZUL       = PatternFill("solid", fgColor="2E75B6")
_F_AZUL_CLARO = PatternFill("solid", fgColor="BDD7EE")
_F_GRIS       = PatternFill("solid", fgColor="D6DCE4")
_F_ROJO       = PatternFill("solid", fgColor="FFE0E0")
_F_VERDE      = PatternFill("solid", fgColor="E2EFDA")
_BLANCO_B     = Font(bold=True, color="FFFFFF", size=9)
_NEGRO_B      = Font(bold=True, size=9)
_NORMAL       = Font(size=9)
_CENTER       = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT         = Alignment(horizontal="left",   vertical="center", wrap_text=True)
_BORDE        = Border(left=Side(style="thin"), right=Side(style="thin"),
                       top=Side(style="thin"),  bottom=Side(style="thin"))

def _cell(ws, row, col, value, fill=None, font=None, align=None):
    c = ws.cell(row=row, column=col, value=value)
    c.border = _BORDE
    if fill:  c.fill  = fill
    if font:  c.font  = font
    c.alignment = align if align else _LEFT
    return c


def _hoja_tabla01(wb, nombre, por_cc, mes, mes_nombre):
    ws = wb.create_sheet(nombre); fila = 1
    for bloque in por_cc:
        cc = bloque["cc"]; acts = bloque["actividades"]
        for c, txt in enumerate([CC_ALIAS.get(cc, cc),"ACTIVIDAD OPERATIVA","U.M.",
                                   "PROG. FÍSICA ANUAL","PROG. FINANCIERA ANUAL",
                                   "PROG. ENE","EJEC. ENE","% ENE",
                                   f"PROG. ENE-{mes_nombre}",f"EJEC. ENE-{mes_nombre}","% ACUM",
                                   "PIM",f"DEVENGADO {mes_nombre}","% FIN."], 1):
            _cell(ws, fila, c+1, txt, _F_AZUL_OSC, _BLANCO_B, _CENTER)
        ws.row_dimensions[fila].height = 35; fila += 1
        for d in acts:
            pct_f = d["pctAvance"]
            fill  = _F_ROJO if (pct_f is not None and pct_f < 75) else (_F_VERDE if (pct_f or 0) >= 95 else None)
            vals  = [None, d["actividadOp"], d["unidadMedida"],
                     d["progAnual"], d["progFinAnual"],
                     d["progMes1"], d["ejecMes1"],
                     f"{d['pctMes1']:.1f}%" if d["pctMes1"] is not None else "—",
                     d["progAcum"], d["ejecAcum"],
                     f"{pct_f:.1f}%" if pct_f is not None else "—",
                     d["pim"], d["devengado"],
                     f"{d['pctFinanciero']:.1f}%" if d["pctFinanciero"] is not None else "—"]
            for c, v in enumerate(vals, 1): _cell(ws, fila, c, v, fill, _NORMAL)
            ws.row_dimensions[fila].height = 18; fila += 1
        fila += 1
    for c, w in enumerate([10,50,8,12,14,12,12,9,12,12,9,14,14,9], 1):
        ws.column_dimensions[get_column_letter(c)].width = w


def _hoja_pp117_entidad(wb, nombre, categorias, mes, mes_nombre):
    ws = wb.create_sheet(nombre); fila = 1
    hdrs = ["CC","PP / PRODUCTO / ACTIVIDAD","U.M."]
    hdrs += [MESES_NOMBRES[i] for i in range(mes)]
    hdrs += [f"PROG. ENE-{mes_nombre}","TOTAL PROG. ANUAL",f"EJEC. ENE-{mes_nombre}",
             "% AVANCE","PIM","DEVENGADO","% EJEC."]
    for c, h in enumerate(hdrs, 1): _cell(ws, fila, c, h, _F_AZUL_OSC, _BLANCO_B, _CENTER)
    ws.row_dimensions[fila].height = 35; fila += 1
    for cat in categorias:
        cat_txt = f"{cat['categoriaId']}.{cat['categoria']}" if cat["categoria"] else cat["categoriaId"]
        _cell(ws, fila, 1, cat_txt, _F_AZUL, _BLANCO_B)
        ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=len(hdrs))
        ws.row_dimensions[fila].height = 20; fila += 1
        for ap in cat["aps"]:
            for d in ap["centros"]:
                alias = CC_ALIAS.get(d["centroCosto"], d["centroCosto"])
                ap_txt = f"{d['actPresupId']}.{d['actPresup']}" if d["actPresup"] else d["actPresupId"]
                pct_f = d["pctAvance"]
                fill  = _F_ROJO if (pct_f is not None and pct_f < 75) else None
                vals  = [alias, ap_txt, d["unidadMedida"]]
                vals += [d["progMeses"][i] for i in range(mes)]
                vals += [d["progAcum"], d["progAnual"], d["ejecAcum"],
                         f"{pct_f:.1f}%" if pct_f is not None else "—",
                         d["pim"], d["devengado"],
                         f"{d['pctFinanciero']:.1f}%" if d["pctFinanciero"] is not None else "—"]
                for c, v in enumerate(vals, 1): _cell(ws, fila, c, v, fill, _NORMAL)
                ws.row_dimensions[fila].height = 16; fila += 1
        fila += 1
    for c, w in enumerate([12,50,8]+[8]*mes+[12,14,12,9,14,14,9], 1):
        ws.column_dimensions[get_column_letter(c)].width = w


# ─── GET /descargar/dgnna ────────────────────────────────────────

@router.get("/descargar/dgnna")
def descargar_dgnna(
    mes: int = Query(..., ge=1, le=12),
    anio: int = Query(..., ge=2020),
    db: Session = Depends(get_db),
    _: dict = Depends(_get_usuario),
):
    carga = _get_carga_o_404(db, mes, anio); mn = MESES_NOMBRES[mes-1]
    filas = db.query(PoiDato).filter(PoiDato.cargaId == carga.id,
                                      PoiDato.centroCosto.in_(CC_DGNNA)).all()
    datos = _agregar_filas(filas, mes)
    cc_grupos: dict = defaultdict(list)
    for d in datos: cc_grupos[d["centroCosto"]].append(d)
    por_cc = [{"cc": cc, "actividades": acts} for cc, acts in cc_grupos.items()]
    wb = Workbook(); wb.remove(wb.active)
    _hoja_tabla01(wb, "TABLA 01 DGNNA", por_cc, mes, mn)
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="Reporte-POI-DGNNA-{mn}{anio}.xlsx"'})


# ─── GET /descargar/pp117 ────────────────────────────────────────

@router.get("/descargar/pp117")
def descargar_pp117(
    mes: int = Query(..., ge=1, le=12),
    anio: int = Query(..., ge=2020),
    db: Session = Depends(get_db),
    _: dict = Depends(_get_usuario),
):
    carga = _get_carga_o_404(db, mes, anio); mn = MESES_NOMBRES[mes-1]

    def _get(cc_f):
        rows = db.query(PoiDato).filter(PoiDato.cargaId == carga.id,
                                         PoiDato.centroCosto.in_(cc_f)).all()
        datos = _agregar_filas(rows, mes)
        cat_map: dict = defaultdict(lambda: {"categoriaId":"","categoria":"",
                        "aps": defaultdict(lambda: {"actPresupId":"","actPresup":"","centros":[]})})
        for d in datos:
            cat = cat_map[d["categoriaId"] or "S/C"]
            cat["categoriaId"] = d["categoriaId"] or ""; cat["categoria"] = d["categoria"] or ""
            ap = cat["aps"][d["actPresupId"] or "S/AP"]
            ap["actPresupId"] = d["actPresupId"] or ""; ap["actPresup"] = d["actPresup"] or ""
            ap["centros"].append(d)
        return [{"categoriaId": c["categoriaId"], "categoria": c["categoria"],
                 "aps": [{"actPresupId": a["actPresupId"], "actPresup": a["actPresup"],
                           "centros": a["centros"]} for a in c["aps"].values()]}
                for c in cat_map.values()]

    wb = Workbook(); wb.remove(wb.active)
    _hoja_pp117_entidad(wb, "PP117-SECTOR MIMP", _get(CC_PP117),   mes, mn)
    _hoja_pp117_entidad(wb, "PP117-DGNNA",       _get(CC_DGNNA),   mes, mn)
    _hoja_pp117_entidad(wb, "PP117-INABIF",      _get(CC_INABIF),  mes, mn)
    _hoja_pp117_entidad(wb, "PP117-CONADIS",     _get(CC_CONADIS), mes, mn)
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="Reporte-PP0117-{mn}{anio}.xlsx"'})
