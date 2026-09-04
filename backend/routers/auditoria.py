"""
Router FastAPI para el módulo de Auditoría y Trazabilidad (Monolito local).
Prefix: /api/auditoria
"""
import io
import json
from datetime import datetime, timedelta
from typing import Optional, List, Dict, Any
from fastapi import APIRouter, Depends, HTTPException, Query, Response
from pydantic import BaseModel, Field
from sqlalchemy.orm import Session
from sqlalchemy import func, desc, or_
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from database import get_db
from models import AuditoriaSistema

router = APIRouter(prefix="/api/auditoria", tags=["auditoria"])


class AuditoriaEventoCreate(BaseModel):
    modulo: str = Field(...)
    tablaAfectada: str = Field(...)
    registroId: str = Field(...)
    codigoReferencia: Optional[str] = None
    accion: str = Field(...)
    camposCambiados: Optional[str] = None
    valoresPrevios: Optional[str] = None
    valoresNuevos: Optional[str] = None
    usuarioId: str = Field(...)
    usuarioNombre: str = Field(...)
    usuarioRol: Optional[str] = None
    ipOrigen: Optional[str] = None


class DiffItem(BaseModel):
    campo: str
    antes: Optional[Any] = None
    despues: Optional[Any] = None


class AuditoriaLogOut(BaseModel):
    id: str
    modulo: str
    tablaAfectada: str
    registroId: str
    codigoReferencia: Optional[str] = None
    accion: str
    camposCambiados: Optional[str] = None
    valoresPrevios: Optional[str] = None
    valoresNuevos: Optional[str] = None
    usuarioId: str
    usuarioNombre: str
    usuarioRol: Optional[str] = None
    ipOrigen: Optional[str] = None
    createdAt: datetime
    diffs: Optional[List[DiffItem]] = []

    class Config:
        from_attributes = True


class AuditoriaPaginadaOut(BaseModel):
    total: int
    page: int
    limit: int
    items: List[AuditoriaLogOut]


class AuditoriaKPIsOut(BaseModel):
    totalHoy: int
    totalModificaciones: int
    totalCreaciones: int
    totalEliminaciones: int
    totalSeguridad: int
    totalGeneral: int
    porModulo: Dict[str, int]


def _parse_diffs(valores_previos_str: Optional[str], valores_nuevos_str: Optional[str]) -> List[DiffItem]:
    diffs: List[DiffItem] = []
    previos = {}
    nuevos = {}
    if valores_previos_str:
        try:
            previos = json.loads(valores_previos_str)
        except Exception:
            pass
    if valores_nuevos_str:
        try:
            nuevos = json.loads(valores_nuevos_str)
        except Exception:
            pass

    todos_los_campos = set(list(previos.keys()) + list(nuevos.keys()))
    for campo in sorted(todos_los_campos):
        val_ant = previos.get(campo)
        val_nue = nuevos.get(campo)
        if val_ant != val_nue:
            diffs.append(DiffItem(
                campo=campo,
                antes=val_ant,
                despues=val_nue
            ))
    return diffs


# ─── GET /api/auditoria ──────────────────────────────────────────
@router.get("", response_model=AuditoriaPaginadaOut)
def listar_auditoria(
    modulo: Optional[str] = Query(None),
    accion: Optional[str] = Query(None),
    usuario: Optional[str] = Query(None),
    registroId: Optional[str] = Query(None),
    busqueda: Optional[str] = Query(None),
    fechaDesde: Optional[str] = Query(None),
    fechaHasta: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    limit: int = Query(25, ge=1, le=200),
    db: Session = Depends(get_db)
):
    query = db.query(AuditoriaSistema)

    if modulo and modulo.lower() != "todos":
        query = query.filter(AuditoriaSistema.modulo == modulo.lower())
    if accion and accion.lower() != "todas":
        query = query.filter(AuditoriaSistema.accion == accion.upper())
    if usuario:
        query = query.filter(
            or_(
                AuditoriaSistema.usuarioNombre.ilike(f"%{usuario}%"),
                AuditoriaSistema.usuarioId == usuario
            )
        )
    if registroId:
        query = query.filter(AuditoriaSistema.registroId == registroId)
    if busqueda:
        b_clean = f"%{busqueda.strip()}%"
        query = query.filter(
            or_(
                AuditoriaSistema.codigoReferencia.ilike(b_clean),
                AuditoriaSistema.camposCambiados.ilike(b_clean),
                AuditoriaSistema.usuarioNombre.ilike(b_clean)
            )
        )
    if fechaDesde:
        try:
            fd = datetime.strptime(fechaDesde, "%Y-%m-%d")
            query = query.filter(AuditoriaSistema.createdAt >= fd)
        except ValueError:
            pass
    if fechaHasta:
        try:
            fh = datetime.strptime(fechaHasta, "%Y-%m-%d") + timedelta(days=1)
            query = query.filter(AuditoriaSistema.createdAt < fh)
        except ValueError:
            pass

    total = query.count()
    offset = (page - 1) * limit
    logs = query.order_by(desc(AuditoriaSistema.createdAt)).offset(offset).limit(limit).all()

    items: List[AuditoriaLogOut] = []
    for log in logs:
        diff_list = _parse_diffs(log.valoresPrevios, log.valoresNuevos)
        items.append(AuditoriaLogOut(
            id=log.id,
            modulo=log.modulo,
            tablaAfectada=log.tablaAfectada,
            registroId=log.registroId,
            codigoReferencia=log.codigoReferencia,
            accion=log.accion,
            camposCambiados=log.camposCambiados,
            valoresPrevios=log.valoresPrevios,
            valoresNuevos=log.valoresNuevos,
            usuarioId=log.usuarioId,
            usuarioNombre=log.usuarioNombre,
            usuarioRol=log.usuarioRol,
            ipOrigen=log.ipOrigen,
            createdAt=log.createdAt,
            diffs=diff_list
        ))

    return AuditoriaPaginadaOut(
        total=total,
        page=page,
        limit=limit,
        items=items
    )


# ─── GET /api/auditoria/stats ────────────────────────────────────
@router.get("/stats", response_model=AuditoriaKPIsOut)
def obtener_estadisticas(db: Session = Depends(get_db)):
    hoy_inicio = datetime.utcnow().replace(hour=0, minute=0, second=0, microsecond=0)
    
    total_general = db.query(AuditoriaSistema).count()
    total_hoy = db.query(AuditoriaSistema).filter(AuditoriaSistema.createdAt >= hoy_inicio).count()
    
    total_mod = db.query(AuditoriaSistema).filter(AuditoriaSistema.accion == "MODIFICAR").count()
    total_crea = db.query(AuditoriaSistema).filter(AuditoriaSistema.accion == "CREAR").count()
    total_elim = db.query(AuditoriaSistema).filter(AuditoriaSistema.accion == "ELIMINAR").count()
    total_seg = db.query(AuditoriaSistema).filter(AuditoriaSistema.accion.in_(["LOGIN", "PERMISOS"])).count()

    modulos_query = db.query(
        AuditoriaSistema.modulo,
        func.count(AuditoriaSistema.id)
    ).group_by(AuditoriaSistema.modulo).all()

    por_modulo = {m: c for m, c in modulos_query}

    return AuditoriaKPIsOut(
        totalHoy=total_hoy,
        totalModificaciones=total_mod,
        totalCreaciones=total_crea,
        totalEliminaciones=total_elim,
        totalSeguridad=total_seg,
        totalGeneral=total_general,
        porModulo=por_modulo
    )


# ─── GET /api/auditoria/{id} ─────────────────────────────────────
@router.get("/{id}", response_model=AuditoriaLogOut)
def obtener_detalle_auditoria(id: str, db: Session = Depends(get_db)):
    log = db.query(AuditoriaSistema).filter(AuditoriaSistema.id == id).first()
    if not log:
        raise HTTPException(status_code=404, detail="Registro de auditoría no encontrado")

    diff_list = _parse_diffs(log.valoresPrevios, log.valoresNuevos)
    return AuditoriaLogOut(
        id=log.id,
        modulo=log.modulo,
        tablaAfectada=log.tablaAfectada,
        registroId=log.registroId,
        codigoReferencia=log.codigoReferencia,
        accion=log.accion,
        camposCambiados=log.camposCambiados,
        valoresPrevios=log.valoresPrevios,
        valoresNuevos=log.valoresNuevos,
        usuarioId=log.usuarioId,
        usuarioNombre=log.usuarioNombre,
        usuarioRol=log.usuarioRol,
        ipOrigen=log.ipOrigen,
        createdAt=log.createdAt,
        diffs=diff_list
    )


# ─── POST /api/auditoria ─────────────────────────────────────────
@router.post("", response_model=AuditoriaLogOut, status_code=201)
def registrar_evento_auditoria(body: AuditoriaEventoCreate, db: Session = Depends(get_db)):
    log = AuditoriaSistema(
        modulo=body.modulo.lower().strip(),
        tablaAfectada=body.tablaAfectada.lower().strip(),
        registroId=str(body.registroId),
        codigoReferencia=body.codigoReferencia,
        accion=body.accion.upper().strip(),
        camposCambiados=body.camposCambiados,
        valoresPrevios=body.valoresPrevios,
        valoresNuevos=body.valoresNuevos,
        usuarioId=body.usuarioId,
        usuarioNombre=body.usuarioNombre,
        usuarioRol=body.usuarioRol,
        ipOrigen=body.ipOrigen,
        createdAt=datetime.utcnow()
    )
    db.add(log)
    db.commit()
    db.refresh(log)

    diff_list = _parse_diffs(log.valoresPrevios, log.valoresNuevos)
    return AuditoriaLogOut(
        id=log.id,
        modulo=log.modulo,
        tablaAfectada=log.tablaAfectada,
        registroId=log.registroId,
        codigoReferencia=log.codigoReferencia,
        accion=log.accion,
        camposCambiados=log.camposCambiados,
        valoresPrevios=log.valoresPrevios,
        valoresNuevos=log.valoresNuevos,
        usuarioId=log.usuarioId,
        usuarioNombre=log.usuarioNombre,
        usuarioRol=log.usuarioRol,
        ipOrigen=log.ipOrigen,
        createdAt=log.createdAt,
        diffs=diff_list
    )


# ─── GET /api/auditoria/exportar/excel ───────────────────────────
@router.get("/exportar/excel")
def exportar_auditoria_excel(
    modulo: Optional[str] = Query(None),
    accion: Optional[str] = Query(None),
    fechaDesde: Optional[str] = Query(None),
    fechaHasta: Optional[str] = Query(None),
    db: Session = Depends(get_db)
):
    query = db.query(AuditoriaSistema)

    if modulo and modulo.lower() != "todos":
        query = query.filter(AuditoriaSistema.modulo == modulo.lower())
    if accion and accion.lower() != "todas":
        query = query.filter(AuditoriaSistema.accion == accion.upper())
    if fechaDesde:
        try:
            fd = datetime.strptime(fechaDesde, "%Y-%m-%d")
            query = query.filter(AuditoriaSistema.createdAt >= fd)
        except ValueError:
            pass
    if fechaHasta:
        try:
            fh = datetime.strptime(fechaHasta, "%Y-%m-%d") + timedelta(days=1)
            query = query.filter(AuditoriaSistema.createdAt < fh)
        except ValueError:
            pass

    logs = query.order_by(desc(AuditoriaSistema.createdAt)).limit(1000).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Auditoría DGNNA"

    headers = [
        "Fecha / Hora",
        "Módulo",
        "Código Referencia",
        "Acción",
        "Campos Modificados",
        "Usuario que Registró",
        "Rol",
        "IP Origen"
    ]
    ws.append(headers)

    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

    for col_idx, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for log in logs:
        fecha_str = log.createdAt.strftime("%d/%m/%Y %H:%M:%S") if log.createdAt else ""
        ws.append([
            fecha_str,
            log.modulo.upper(),
            log.codigoReferencia or log.registroId,
            log.accion,
            log.camposCambiados or "-",
            log.usuarioNombre,
            log.usuarioRol or "-",
            log.ipOrigen or "-"
        ])

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    nombre_archivo = f"auditoria_dgnna_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={nombre_archivo}"}
    )
