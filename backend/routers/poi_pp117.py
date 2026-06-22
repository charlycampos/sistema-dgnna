"""
Router: Módulo POI - PP117
Carga mensual del archivo DATA-POI y generación de reportes DGNNA y PP0117.
"""

import io
import math
from collections import defaultdict
from datetime import datetime
from typing import List, Optional

import pandas as pd
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from database import get_db
from models import PoiCarga, PoiDato
from auth import require_module_access, can_write

router = APIRouter(prefix="/api/poi-pp117", tags=["poi-pp117"])
MODULO = "poi-pp117"

MESES_NOMBRES = ["ENE","FEB","MAR","ABR","MAY","JUN","JUL","AGO","SET","OCT","NOV","DIC"]

CC_DGNNA = {
    "DIRECCIÓN DE ADOPCIONES",
    "DIRECCIÓN DE POLÍTICAS DE NIÑAS, NIÑOS Y ADOLESCENTES",
    "DIRECCIÓN DE PROTECCIÓN ESPECIAL",
    "DIRECCIÓN DE SISTEMAS LOCALES Y DEFENSORÍAS",
    "DIRECCIÓN GENERAL DE NIÑAS, NIÑOS Y ADOLESCENTES",
}
CC_INABIF = {
    "UNIDAD DE ADMINISTRACIÓN",
    "UNIDAD DE DESARROLLO INTEGRAL DE LAS FAMILIAS",
    "UNIDAD DE FORTALECIMIENTO DE SERVICIOS Y COORDINACION TERRITORIAL",
    "UNIDAD DE SERVICIOS DE PROTECCIÓN DE NIÑOS, NIÑAS Y ADOLESCENTES",
    "UNIDAD DE SERVICIOS TERRITORIALES",
}
CC_CONADIS = {"CONADIS"}
CC_PP117   = CC_DGNNA | CC_INABIF | CC_CONADIS

# Abreviaturas para mostrar en gráficos
CC_ALIAS = {
    "DIRECCIÓN DE ADOPCIONES": "DA",
    "DIRECCIÓN DE POLÍTICAS DE NIÑAS, NIÑOS Y ADOLESCENTES": "DPNNA",
    "DIRECCIÓN DE PROTECCIÓN ESPECIAL": "DPE",
    "DIRECCIÓN DE SISTEMAS LOCALES Y DEFENSORÍAS": "DSLD",
    "DIRECCIÓN GENERAL DE NIÑAS, NIÑOS Y ADOLESCENTES": "DGNNA",
    "UNIDAD DE ADMINISTRACIÓN": "UA",
    "UNIDAD DE DESARROLLO INTEGRAL DE LAS FAMILIAS": "UDIF",
    "UNIDAD DE FORTALECIMIENTO DE SERVICIOS Y COORDINACION TERRITORIAL": "UFSCT",
    "UNIDAD DE SERVICIOS DE PROTECCIÓN DE NIÑOS, NIÑAS Y ADOLESCENTES": "USPNNA",
    "UNIDAD DE SERVICIOS TERRITORIALES": "UST",
    "CONADIS": "CONADIS",
}

# ─── Utilidades ──────────────────────────────────────────────────

def _safe(val) -> Optional[float]:
    if val is None: return None
    try:
        v = float(val)
        return None if math.isnan(v) else v
    except (TypeError, ValueError):
        return None

def _pct(ejec, prog) -> Optional[float]:
    if prog and prog > 0 and ejec is not None:
        return round(ejec / prog * 100, 1)
    return None

def _str(val) -> Optional[str]:
    s = str(val or "").strip()
    return s or None

def _get_carga_o_404(db: Session, mes: int, anio: int) -> PoiCarga:
    c = db.query(PoiCarga).filter(PoiCarga.mes == mes, PoiCarga.anio == anio).first()
    if not c:
        raise HTTPException(404, f"No hay datos para {MESES_NOMBRES[mes-1]} {anio}")
    return c


# ─── GET /meses ──────────────────────────────────────────────────

@router.get("/meses")
def listar_meses(db: Session = Depends(get_db), _: dict = Depends(require_module_access(MODULO))):
    cargas = db.query(PoiCarga).order_by(PoiCarga.anio.desc(), PoiCarga.mes.desc()).all()
    return [
        {"id": c.id, "mes": c.mes, "anio": c.anio,
         "mesNombre": MESES_NOMBRES[c.mes - 1],
         "nombreArchivo": c.nombreArchivo, "totalFilas": c.totalFilas,
         "cargadoPor": c.cargadoPor,
         "createdAt": c.createdAt.isoformat() if c.createdAt else None}
        for c in cargas
    ]


# ─── POST /cargar ────────────────────────────────────────────────

@router.post("/cargar", status_code=201)
async def cargar_archivo(
    mes: int = Query(..., ge=1, le=12),
    anio: int = Query(..., ge=2020, le=2099),
    reemplazar: bool = Query(False),
    archivo: UploadFile = File(...),
    db: Session = Depends(get_db),
    current: dict = Depends(can_write(MODULO)),
):
    existente = db.query(PoiCarga).filter(PoiCarga.mes == mes, PoiCarga.anio == anio).first()
    if existente and not reemplazar:
        raise HTTPException(409, f"Ya existe carga para {MESES_NOMBRES[mes-1]} {anio}. Use reemplazar=true.")

    contenido = await archivo.read()
    try:
        df = pd.read_excel(io.BytesIO(contenido), sheet_name="POI_Por_ActividadOperativaAnual")
    except Exception as e:
        raise HTTPException(400, f"Error leyendo Excel: {e}")

    if existente:
        db.delete(existente); db.commit()

    carga = PoiCarga(mes=mes, anio=anio, nombreArchivo=archivo.filename,
                     totalFilas=len(df), cargadoPor=current.get("nombre"))
    db.add(carga); db.flush()

    mes_motivo_col = f"Motivo(SE) {mes:02d}"
    registros = []

    for _, row in df.iterrows():
        cc = _str(row.get("Centro de Costo")) or ""
        if cc not in CC_PP117:
            continue

        dato = PoiDato(
            cargaId      = carga.id,
            centroCosto  = cc,
            categoriaId  = _str(row.get("Categoria ID")),
            categoria    = _str(row.get("Categoria")),
            actPresupId  = _str(row.get("Actividad Presupuestal ID")),
            actPresup    = _str(row.get("Actividad Presupuestal")),
            codAO        = _str(row.get("Actividad Operativa ID")),
            actividadOp  = _str(row.get("Actividad Operativa")),
            unidadMedida = _str(row.get("Unidad de Medida")),
            departamento = _str(row.get("Departamento Nombre UBIGEO")),
            consPIM      = _safe(row.get("Cons PIM")),
            fnReTotal    = _safe(row.get("Fn(RE) Total")),
            fRe01=_safe(row.get("F(RE) 01")), fRe02=_safe(row.get("F(RE) 02")),
            fRe03=_safe(row.get("F(RE) 03")), fRe04=_safe(row.get("F(RE) 04")),
            fRe05=_safe(row.get("F(RE) 05")), fRe06=_safe(row.get("F(RE) 06")),
            fRe07=_safe(row.get("F(RE) 07")), fRe08=_safe(row.get("F(RE) 08")),
            fRe09=_safe(row.get("F(RE) 09")), fRe10=_safe(row.get("F(RE) 10")),
            fRe11=_safe(row.get("F(RE) 11")), fRe12=_safe(row.get("F(RE) 12")),
            fReTotal=_safe(row.get("F(RE) Total")),
            fSe01=_safe(row.get("F(SE) 01")), fSe02=_safe(row.get("F(SE) 02")),
            fSe03=_safe(row.get("F(SE) 03")), fSe04=_safe(row.get("F(SE) 04")),
            fSe05=_safe(row.get("F(SE) 05")), fSe06=_safe(row.get("F(SE) 06")),
            fSe07=_safe(row.get("F(SE) 07")), fSe08=_safe(row.get("F(SE) 08")),
            fSe09=_safe(row.get("F(SE) 09")), fSe10=_safe(row.get("F(SE) 10")),
            fSe11=_safe(row.get("F(SE) 11")), fSe12=_safe(row.get("F(SE) 12")),
            fSeTotal=_safe(row.get("F(SE) Total")),
            fnSeTotal=_safe(row.get("Fn(SE) Total")),
            motivoMes=_str(row.get(mes_motivo_col)) if mes_motivo_col in df.columns else None,
        )
        registros.append(dato)

    db.bulk_save_objects(registros); db.commit()
    return {"mensaje": f"Carga exitosa: {MESES_NOMBRES[mes-1]} {anio}",
            "filasGuardadas": len(registros), "totalFilas": len(df)}


# ─── Agregación central ──────────────────────────────────────────

def _agregar_filas(filas: List[PoiDato], mes_eval: int) -> List[dict]:
    """Agrega filas al nivel CC+AO (suma departamentos)."""
    grupos: dict = defaultdict(lambda: {
        "centroCosto":"","categoriaId":"","categoria":"",
        "actPresupId":"","actPresup":"",
        "codAO":"","actividadOp":"","unidadMedida":"","motivoMes":"",
        "fRe":[0.0]*12,"fReTotal":0.0,
        "fSe":[0.0]*12,"fSeTotal":0.0,
        "consPIM":0.0,"fnReTotal":0.0,"fnSeTotal":0.0,
    })

    re_attrs = [f"fRe{i:02d}" for i in range(1,13)]
    se_attrs = [f"fSe{i:02d}" for i in range(1,13)]

    for f in filas:
        key = (f.centroCosto, f.codAO)
        g = grupos[key]
        g["centroCosto"]  = f.centroCosto or ""
        g["categoriaId"]  = f.categoriaId or ""
        g["categoria"]    = f.categoria   or g["categoria"]
        g["actPresupId"]  = f.actPresupId or ""
        g["actPresup"]    = f.actPresup   or g["actPresup"]
        g["codAO"]        = f.codAO       or ""
        g["actividadOp"]  = f.actividadOp or ""
        g["unidadMedida"] = f.unidadMedida or ""
        g["motivoMes"]    = f.motivoMes   or g["motivoMes"]
        for i, a in enumerate(re_attrs): g["fRe"][i] += (getattr(f,a) or 0)
        g["fReTotal"]  += (f.fReTotal  or 0)
        for i, a in enumerate(se_attrs): g["fSe"][i] += (getattr(f,a) or 0)
        g["fSeTotal"]  += (f.fSeTotal  or 0)
        g["consPIM"]   += (f.consPIM   or 0)
        g["fnReTotal"] += (f.fnReTotal or 0)
        g["fnSeTotal"] += (f.fnSeTotal or 0)

    resultado = []
    for g in grupos.values():
        prog_acum = sum(g["fRe"][:mes_eval])
        ejec_acum = sum(g["fSe"][:mes_eval])
        prog_mes1 = g["fRe"][0]
        ejec_mes1 = g["fSe"][0]
        resultado.append({
            "centroCosto":   g["centroCosto"],
            "ccAlias":       CC_ALIAS.get(g["centroCosto"], g["centroCosto"][:6]),
            "categoriaId":   g["categoriaId"],
            "categoria":     g["categoria"],
            "actPresupId":   g["actPresupId"],
            "actPresup":     g["actPresup"],
            "codAO":         g["codAO"],
            "actividadOp":   g["actividadOp"],
            "unidadMedida":  g["unidadMedida"],
            "progAnual":     round(g["fReTotal"], 2),
            "progFinAnual":  round(g["fnReTotal"], 2),
            "progMeses":     [round(v,2) for v in g["fRe"]],
            "ejecMeses":     [round(v,2) for v in g["fSe"]],
            "progMes1":      round(prog_mes1, 2),
            "ejecMes1":      round(ejec_mes1, 2),
            "pctMes1":       _pct(ejec_mes1, prog_mes1),
            "progAcum":      round(prog_acum, 2),
            "ejecAcum":      round(ejec_acum, 2),
            "pctAvance":     _pct(ejec_acum, prog_acum),
            "pim":           round(g["consPIM"], 2),
            "devengado":     round(g["fnSeTotal"], 2),
            "pctFinanciero": _pct(g["fnSeTotal"], g["consPIM"]),
            "motivoMes":     g["motivoMes"],
        })
    return resultado


# ─── GET /dashboard ──────────────────────────────────────────────

@router.get("/dashboard")
def dashboard(
    mes: int   = Query(..., ge=1, le=12),
    anio: int  = Query(..., ge=2020),
    tipo: str  = Query("dgnna"),   # dgnna | pp117
    db: Session = Depends(get_db),
    _: dict     = Depends(require_module_access(MODULO)),
):
    carga = _get_carga_o_404(db, mes, anio)
    cc_filtro = CC_DGNNA if tipo == "dgnna" else CC_PP117

    filas = db.query(PoiDato).filter(
        PoiDato.cargaId == carga.id,
        PoiDato.centroCosto.in_(cc_filtro),
    ).all()
    datos = _agregar_filas(filas, mes)

    # KPIs globales
    prog_total  = sum(d["progAcum"]  for d in datos)
    ejec_total  = sum(d["ejecAcum"]  for d in datos)
    pim_total   = sum(d["pim"]       for d in datos)
    dev_total   = sum(d["devengado"] for d in datos)

    # Semáforo
    verde    = sum(1 for d in datos if (d["pctAvance"] or 0) >= 80)
    amarillo = sum(1 for d in datos if 50 <= (d["pctAvance"] or 0) < 80)
    rojo     = sum(1 for d in datos if d["pctAvance"] is not None and d["pctAvance"] < 50)

    # Por centro de costo
    cc_map: dict = defaultdict(lambda: {"alias":"","prog":0,"ejec":0,"pim":0,"dev":0})
    for d in datos:
        cc = d["centroCosto"]
        cc_map[cc]["alias"] = d["ccAlias"]
        cc_map[cc]["prog"]  += d["progAcum"]
        cc_map[cc]["ejec"]  += d["ejecAcum"]
        cc_map[cc]["pim"]   += d["pim"]
        cc_map[cc]["dev"]   += d["devengado"]

    por_cc = [
        {"cc": alias, "ccCompleto": cc,
         "pctFisico":     round(v["ejec"]/v["prog"]*100, 1) if v["prog"] else 0,
         "pctFinanciero": round(v["dev"]/v["pim"]*100, 1)   if v["pim"]  else 0,
         "ejecAcum": round(v["ejec"], 2), "progAcum": round(v["prog"], 2)}
        for cc, v in cc_map.items()
        for alias in [CC_ALIAS.get(cc, cc[:8])]
    ]
    por_cc.sort(key=lambda x: x["pctFisico"], reverse=True)

    # Mensual (prog vs ejec sumados)
    mensual = [
        {"mes": MESES_NOMBRES[i],
         "prog": round(sum(d["progMeses"][i] for d in datos), 2),
         "ejec": round(sum(d["ejecMeses"][i] for d in datos), 2)}
        for i in range(mes)
    ]

    return {
        "mes": mes, "anio": anio, "mesNombre": MESES_NOMBRES[mes-1],
        "kpis": {
            "totalAOs":       len(datos),
            "pctAvanceFisico": _pct(ejec_total, prog_total) or 0,
            "pimTotal":        round(pim_total, 2),
            "devengadoTotal":  round(dev_total, 2),
            "pctFinanciero":   _pct(dev_total, pim_total) or 0,
        },
        "semaforo": {"verde": verde, "amarillo": amarillo, "rojo": rojo},
        "porCC":    por_cc,
        "mensual":  mensual,
    }


TABLA01_AOS_PERMITIDAS = {
    "DA": [
        "Evaluación de la capacidad a las familias solicitantes de adopción",
        "Integración familiar de la niña, niño o adolescente declarado judicialmente en desprotección familiar y adoptabilidad",
        "Acompañamiento post adoptivo a la niña, niño o adolescente y a su familia",
        "Gestión del Programa"
    ],
    "DSLD": [
        "Intervención lúdica y espacios seguros para el fortalecimiento de capacidades de niñas, niños y adolescentes- Multisectorial",
        "Intervención lúdica y espacios seguros para el fortalecimiento de capacidades de niñas, niños y adolescentes",
        "Acreditación y supervisión de DEMUNAS",
        "Fortalecimiento de las defensorías municipales de la niña, niño y adolescente – DEMUNA",
        "Fortalecimiento de los Consejos Consultivos y Participación de Niñas, Niños y Adolescentes – CCONNA",
        "Implementación de la estrategia Ponte en #ModoNiñez",
    ],
    "DPNNA": [
        "Supervisión y acreditación de centros de acogida residencial públicos y privados",
        "Implementación seguimiento y evaluación de la Política Nacional Multisectorial para las Niñas, Niños y Adolescentes",
        "Fortalecimiento de los servicios de atención a niñas, niños y adolescentes vulnerables al delito de explotación sexual en el ámbito nacional",
        "Seguimiento a la atención y reintegración a niñas, niños y adolescentes afectados por el delito de trata de personas en el ámbito nacional.",
    ],
    "DGNNA": [
        "Gestión del Programa Presupuestal 0117",
        "Fortalecimiento del sistema de protección de niñas, niños y adolescentes.",
        "Atención de restitución del ejercicio de derechos de niñas, niños y adolescentes",
        "Implementación de la Estrategia Multisectorial para la prevención de la violencia sexual que afectan a niñas, niños y adolescentes \"PREVENIR PARA PROTEGER\"",
    ],
    "UPE": [
        "Atención preliminar de niñas, niños y adolescentes ingresados al servicio",
        "Elaboración y aprobación del Plan de Trabajo Individual de niñas, niños y adolescentes con Acogimiento Familiar",
        "Elaboración y aprobación del Plan de Trabajo Individual de niñas, niños y adolescentes con Acogimiento Residencial",
        "Elaboración y aprobación del Plan de Trabajo Individual de niñas, niños y adolescentes en riesgo de desprotección familiar",
        "Evaluación sociofamiliar de niñas, niños y adolescentes ingresados al servicio",
    ],
    "DPE": [
        "Gestión del Programa",
        "Atención de llamadas a través de la LÍNEA ANNA 1810",
        "Supervisión a las UPE",
        "Evaluación , declaración de capacidad y capacitación de la familia para el Acogimiento Familiar",
        "Subvención económica para Acogimiento Familiar a las Niñas, Niños y Adolescentes",
    ]
}

def _agregar_filas_tabla01(filas, mes_eval):
    import unicodedata
    def normalizar(s):
        if not s: return ""
        s = s.replace("–", "-")
        return unicodedata.normalize('NFKD', s).encode('ASCII', 'ignore').decode('utf-8').upper().strip()

    AO_TYPOS = {
        "ATENCIÓN PRELIMINAR DE NIÑAS, NIÑOS Y ADOELSCENTES INGRESADOS AL SERVICIO": "Atención preliminar de niñas, niños y adolescentes ingresados al servicio",
    }
    
    # Mapeo de nombre de CC en BD a nuestro Alias de Grupo
    DB_CC_A_GRUPO = {
        "DIRECCIÓN DE ADOPCIONES": "DA",
        "DIRECCIÓN DE SISTEMAS LOCALES Y DEFENSORÍAS": "DSLD",
        "DIRECCIÓN DE POLÍTICAS DE NIÑAS, NIÑOS Y ADOLESCENTES": "DPNNA",
        "DIRECCIÓN GENERAL DE NIÑAS, NIÑOS Y ADOLESCENTES": "DGNNA",
    }

    # Para buscar si una AO pertenece a UPE o DPE (que comparten el mismo CC en la BD)
    upe_aos_norm = {normalizar(a): a for a in TABLA01_AOS_PERMITIDAS["UPE"]}
    dpe_aos_norm = {normalizar(a): a for a in TABLA01_AOS_PERMITIDAS["DPE"]}
    
    # Para validar otras AOs
    cc_permitidas_norm = {}
    for cc, acts in TABLA01_AOS_PERMITIDAS.items():
        cc_permitidas_norm[cc] = {normalizar(a): a for a in acts}

    grupos = {}
    for f in filas:
        cc_db = f.centroCosto
        
        ao_nombre = (f.actividadOp or "").strip()
        if " - " in ao_nombre:
            ao_prefijo = ao_nombre.split(" - ", 1)[0].strip()
        else:
            ao_prefijo = ao_nombre
            
        ao_prefijo = AO_TYPOS.get(ao_prefijo, ao_prefijo)
        ao_norm = normalizar(ao_prefijo)
        
        cc_group = None
        ao_oficial = None
        
        if cc_db == "DIRECCIÓN DE PROTECCIÓN ESPECIAL":
            if ao_norm in upe_aos_norm:
                cc_group = "UPE"
                ao_oficial = upe_aos_norm[ao_norm]
            elif ao_norm in dpe_aos_norm:
                cc_group = "DPE"
                ao_oficial = dpe_aos_norm[ao_norm]
            else:
                continue
        else:
            cc_group = DB_CC_A_GRUPO.get(cc_db)
            if not cc_group:
                continue
            if ao_norm in cc_permitidas_norm[cc_group]:
                ao_oficial = cc_permitidas_norm[cc_group][ao_norm]
            else:
                continue
        
        k = (cc_group, ao_oficial)
        if k not in grupos:
            grupos[k] = {
                "centroCosto": cc_group,
                "codAO": "",
                "actividadOp": ao_oficial,
                "unidadMedida": f.unidadMedida or "",
                "motivoMes": "",
                "fRe": [0]*12, "fReTotal": 0,
                "fSe": [0]*12, "fSeTotal": 0,
                "consPIM": 0, "fnReTotal": 0, "fnSeTotal": 0,
            }
        
        g = grupos[k]
        g["unidadMedida"] = g["unidadMedida"] or (f.unidadMedida or "")
        g["motivoMes"]    = f.motivoMes or g["motivoMes"]
        for i in range(1, 13):
            g["fRe"][i-1] += (getattr(f, f"fRe{i:02d}") or 0)
            g["fSe"][i-1] += (getattr(f, f"fSe{i:02d}") or 0)
            
        g["fReTotal"]  += (f.fReTotal  or 0)
        g["fSeTotal"]  += (f.fSeTotal  or 0)
        g["consPIM"]   += (f.consPIM   or 0)
        g["fnReTotal"] += (f.fnReTotal or 0)
        g["fnSeTotal"] += (f.fnSeTotal or 0)
        
    resultado = []
    # Usar nombres completos agradables para la tabla
    ALIAS_A_NOMBRE = {
        "DA": "DIRECCIÓN DE ADOPCIONES",
        "DSLD": "DIRECCIÓN DE SISTEMAS LOCALES Y DEFENSORÍAS",
        "DPNNA": "DIRECCIÓN DE POLÍTICAS DE NIÑAS, NIÑOS Y ADOLESCENTES",
        "DGNNA": "DIRECCIÓN GENERAL DE NIÑAS, NIÑOS Y ADOLESCENTES",
        "UPE": "UNIDAD DE PROTECCIÓN ESPECIAL (UPE)",
        "DPE": "DIRECCIÓN DE PROTECCIÓN ESPECIAL",
    }
    for cc, acts in TABLA01_AOS_PERMITIDAS.items():
        for ao_oficial in acts:
            k = (cc, ao_oficial)
            if k in grupos:
                g = grupos[k]
                prog_acum = sum(g["fRe"][:mes_eval])
                ejec_acum = sum(g["fSe"][:mes_eval])
                resultado.append({
                    "centroCosto":   cc,
                    "ccAlias":       ALIAS_A_NOMBRE.get(cc, cc),
                    "codAO":         g["codAO"],
                    "actividadOp":   g["actividadOp"],
                    "unidadMedida":  g["unidadMedida"],
                    "progAnual":     round(g["fReTotal"], 2),
                    "progFinAnual":  round(g["fnReTotal"], 2),
                    "progAcum":      round(prog_acum, 2),
                    "ejecAcum":      round(ejec_acum, 2),
                    "pctAvance":     _pct(ejec_acum, prog_acum),
                    "pim":           round(g["consPIM"], 2),
                    "devengado":     round(g["fnSeTotal"], 2),
                    "pctFinanciero": _pct(g["fnSeTotal"], g["consPIM"]),
                    "motivoMes":     g["motivoMes"],
                })
    return resultado


# ─── GET /dgnna ──────────────────────────────────────────────────

@router.get("/dgnna")
def datos_dgnna(
    mes: int = Query(..., ge=1, le=12),
    anio: int = Query(..., ge=2020),
    db: Session = Depends(get_db),
    _: dict = Depends(require_module_access(MODULO)),
):
    carga = _get_carga_o_404(db, mes, anio)
    filas = db.query(PoiDato).filter(
        PoiDato.cargaId == carga.id,
        PoiDato.centroCosto.in_(CC_DGNNA),
    ).all()
    datos = _agregar_filas_tabla01(filas, mes)

    # Agrupar por CC para TABLA 01
    cc_grupos: dict = defaultdict(list)
    for d in datos:
        cc_grupos[d["centroCosto"]].append(d)

    return {
        "mes": mes, "anio": anio, "mesNombre": MESES_NOMBRES[mes-1],
        "datos": datos,
        "porCC": [{"cc": cc, "ccAlias": CC_ALIAS.get(cc, cc), "actividades": acts}
                  for cc, acts in cc_grupos.items()],
    }


# ─── GET /dgnna-upe-debug (temporal) ─────────────────────────────
@router.get("/dgnna-upe-debug")
def debug_dgnna_upe(
    mes: int = Query(..., ge=1, le=12),
    anio: int = Query(..., ge=2020),
    db: Session = Depends(get_db),
):
    from sqlalchemy import func, distinct
    carga = _get_carga_o_404(db, mes, anio)
    filas = db.query(PoiDato.actividadOp).filter(
        PoiDato.cargaId == carga.id,
        PoiDato.centroCosto.in_(CC_DGNNA),
    ).distinct().all()
    prefijos = set()
    for (ao,) in filas:
        if ao and " - " in ao:
            prefijos.add(ao.split(" - ", 1)[0].strip())
        elif ao:
            prefijos.add(ao.strip())
    return {"prefijos": sorted(prefijos), "total_filas": len(filas)}


# ─── GET /dgnna-upe ──────────────────────────────────────────────

@router.get("/dgnna-upe")
def datos_dgnna_upe(
    mes: int = Query(..., ge=1, le=12),
    anio: int = Query(..., ge=2020),
    db: Session = Depends(get_db),
    _: dict = Depends(require_module_access(MODULO)),
):
    carga = _get_carga_o_404(db, mes, anio)

    # Los 5 tipos de AO de la TABLA 02-UPE, en orden
    AO_TIPOS = [
        "ATENCIÓN PRELIMINAR DE NIÑAS, NIÑOS Y ADOLESCENTES INGRESADOS AL SERVICIO",
        "ELABORACIÓN Y APROBACIÓN DEL PLAN DE TRABAJO INDIVIDUAL DE NIÑAS, NIÑOS Y ADOLESCENTES CON ACOGIMIENTO FAMILIAR",
        "ELABORACIÓN Y APROBACIÓN DEL PLAN DE TRABAJO INDIVIDUAL DE NIÑAS, NIÑOS Y ADOLESCENTES CON ACOGIMIENTO RESIDENCIAL",
        "ELABORACIÓN Y APROBACIÓN DEL PLAN DE TRABAJO INDIVIDUAL DE NIÑAS, NIÑOS Y ADOLESCENTES EN RIESGO DE DESPROTECCION FAMILIAR",
        "EVALUACIÓN SOCIOFAMILIAR DE NIÑAS, NIÑOS Y ADOLESCENTES INGRESADOS AL SERVICIO",
    ]

    filas = db.query(PoiDato).filter(
        PoiDato.cargaId == carga.id,
        PoiDato.centroCosto.in_(CC_DGNNA),
    ).all()
    # Typos conocidos en el DATA-POI
    AO_TYPOS = {
        # AO1 San Martín con typo "ADOELSCENTES"
        "ATENCIÓN PRELIMINAR DE NIÑAS, NIÑOS Y ADOELSCENTES INGRESADOS AL SERVICIO":
            "ATENCIÓN PRELIMINAR DE NIÑAS, NIÑOS Y ADOLESCENTES INGRESADOS AL SERVICIO",
    }

    ao_grupos: dict = {t: {"codAO": t, "actividadOp": t, "unidadMedida": "", "upes": []} for t in AO_TIPOS}

    for f in filas:
        ao_nombre = (f.actividadOp or "").strip()
        # Separar prefijo (tipo de AO) y sufijo (nombre de UPE)
        if " - " in ao_nombre:
            partes    = ao_nombre.split(" - ", 1)
            ao_prefijo = partes[0].strip()
            upe_sufijo = partes[1].strip()
        else:
            ao_prefijo = ao_nombre
            upe_sufijo = ""
        # Corregir typos conocidos en el prefijo
        ao_prefijo = AO_TYPOS.get(ao_prefijo, ao_prefijo)
        # Solo procesar si es uno de los 5 tipos del reporte
        if ao_prefijo not in ao_grupos:
            continue
        g = ao_grupos[ao_prefijo]
        g["unidadMedida"] = g["unidadMedida"] or (f.unidadMedida or "")
        # Normalizar nombre de UPE: asegurar prefijo "UPE "
        if upe_sufijo.upper().startswith("UPE "):
            upe_nombre = upe_sufijo.upper()
        elif upe_sufijo:
            upe_nombre = f"UPE {upe_sufijo.upper()}"
        else:
            dep = (f.departamento or "").upper()
            upe_nombre = f"UPE {dep}" if dep else "SIN UPE"
        re_vals    = [getattr(f, f"fRe{i:02d}") or 0 for i in range(1, 13)]
        se_vals    = [getattr(f, f"fSe{i:02d}") or 0 for i in range(1, 13)]
        prog_anual = round(f.fReTotal or 0, 2)
        # F(RE) y F(SE) son valores ACUMULADOS al mes: usar el del mes actual
        prog       = re_vals[mes - 1]
        ejec       = se_vals[mes - 1]
        # sinDatos: sin programación física anual ni ejecución física acumulada
        sin_datos  = (prog_anual == 0 and ejec == 0)
        # PIM: Fn(RE) Total (programación financiera reformulada)
        pim       = round(f.fnReTotal or 0, 2)
        devengado = round(f.fnSeTotal or 0, 2)
        g["upes"].append({
            "departamento":   upe_nombre,
            "progAnual":      prog_anual,
            "progAcum":       round(prog, 2),
            "ejecAcum":       round(ejec, 2),
            "pctAvance":      _pct(ejec, prog),
            "pim":            pim,
            "devengado":      devengado,
            "pctFinanciero":  _pct(devengado, pim),
            "metaPptal":      None,  # META PPTAL — dato SIAF, no disponible en POI
            "sinDatos":       sin_datos,
        })

    return {
        "mes": mes, "anio": anio, "mesNombre": MESES_NOMBRES[mes-1],
        "actividades": [v for v in ao_grupos.values() if v["upes"]],
    }


# ─── GET /estructura-pp117 ───────────────────────────────────────

@router.get("/estructura-pp117")
def estructura_pp117(
    mes: int     = Query(..., ge=1, le=12),
    anio: int    = Query(..., ge=2020),
    entidad: str = Query("sector"),   # sector | dgnna | inabif | conadis
    db: Session  = Depends(get_db),
    _: dict      = Depends(require_module_access(MODULO)),
):
    carga = _get_carga_o_404(db, mes, anio)
    cc_map = {"dgnna": CC_DGNNA, "inabif": CC_INABIF,
              "conadis": CC_CONADIS, "sector": CC_PP117}
    cc_filtro = cc_map.get(entidad, CC_PP117)

    filas = db.query(PoiDato).filter(
        PoiDato.cargaId == carga.id,
        PoiDato.centroCosto.in_(cc_filtro),
    ).all()
    datos = _agregar_filas(filas, mes)

    # Agrupar: categoriaId → actPresupId → centros
    cat_map: dict = defaultdict(lambda: {
        "categoriaId":"","categoria":"","totalProg":0,"totalEjec":0,
        "aps": defaultdict(lambda: {"actPresupId":"","actPresup":"","totalProg":0,"totalEjec":0,"centros":[]})
    })

    for d in datos:
        cat_id = d["categoriaId"] or "S/C"
        ap_id  = d["actPresupId"] or "S/AP"
        cat = cat_map[cat_id]
        cat["categoriaId"] = cat_id
        cat["categoria"]   = d["categoria"] or cat_id
        cat["totalProg"]  += d["progAcum"]
        cat["totalEjec"]  += d["ejecAcum"]
        ap = cat["aps"][ap_id]
        ap["actPresupId"] = ap_id
        ap["actPresup"]   = d["actPresup"] or ap_id
        ap["totalProg"]  += d["progAcum"]
        ap["totalEjec"]  += d["ejecAcum"]
        ap["centros"].append(d)

    categorias = []
    for cat in cat_map.values():
        aps = []
        for ap in cat["aps"].values():
            aps.append({
                "actPresupId": ap["actPresupId"],
                "actPresup":   ap["actPresup"],
                "totalProg":   round(ap["totalProg"], 2),
                "totalEjec":   round(ap["totalEjec"], 2),
                "pctAvance":   _pct(ap["totalEjec"], ap["totalProg"]),
                "centros":     ap["centros"],
            })
        categorias.append({
            "categoriaId": cat["categoriaId"],
            "categoria":   cat["categoria"],
            "totalProg":   round(cat["totalProg"], 2),
            "totalEjec":   round(cat["totalEjec"], 2),
            "pctAvance":   _pct(cat["totalEjec"], cat["totalProg"]),
            "aps":         aps,
        })

    return {
        "mes": mes, "anio": anio,
        "mesNombre": MESES_NOMBRES[mes-1],
        "entidad": entidad,
        "categorias": categorias,
    }


# ─── DELETE /carga/{id} ──────────────────────────────────────────

@router.delete("/carga/{id}", status_code=204)
def eliminar_carga(id: str, db: Session = Depends(get_db), _: dict = Depends(can_write(MODULO))):
    c = db.query(PoiCarga).filter(PoiCarga.id == id).first()
    if not c: raise HTTPException(404, "Carga no encontrada")
    db.delete(c); db.commit()


# ─── Helpers Excel ───────────────────────────────────────────────

_F_AZUL_OSC = PatternFill("solid", fgColor="1F4E79")
_F_AZUL     = PatternFill("solid", fgColor="2E75B6")
_F_AZUL_CLARO = PatternFill("solid", fgColor="BDD7EE")
_F_GRIS     = PatternFill("solid", fgColor="D6DCE4")
_F_ROJO     = PatternFill("solid", fgColor="FFE0E0")
_F_VERDE    = PatternFill("solid", fgColor="E2EFDA")
_BLANCO_B   = Font(bold=True, color="FFFFFF", size=9)
_NEGRO_B    = Font(bold=True, size=9)
_NORMAL     = Font(size=9)
_CENTER     = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT       = Alignment(horizontal="left",   vertical="center", wrap_text=True)
_BORDE      = Border(left=Side(style="thin"), right=Side(style="thin"),
                     top=Side(style="thin"),  bottom=Side(style="thin"))

def _cell(ws, row, col, value, fill=None, font=None, align=None):
    c = ws.cell(row=row, column=col, value=value)
    c.border = _BORDE
    if fill:  c.fill  = fill
    if font:  c.font  = font
    if align: c.alignment = align
    else:     c.alignment = _LEFT
    return c


def _hoja_tabla01(wb: Workbook, nombre: str, por_cc: list, mes: int, mes_nombre: str, anio: int):
    ws = wb.create_sheet(nombre)
    fila = 1

    for bloque in por_cc:
        cc      = bloque["cc"]
        alias   = CC_ALIAS.get(cc, cc)
        acts    = bloque["actividades"]

        # Encabezado del bloque
        _cell(ws, fila, 1, alias, _F_AZUL_OSC, _BLANCO_B, _CENTER)
        _cell(ws, fila, 2, "ACTIVIDAD OPERATIVA", _F_AZUL_OSC, _BLANCO_B, _CENTER)
        _cell(ws, fila, 3, "U.M.", _F_AZUL_OSC, _BLANCO_B, _CENTER)
        _cell(ws, fila, 4, "PROG.\nFÍSICA\nANUAL", _F_AZUL_OSC, _BLANCO_B, _CENTER)
        _cell(ws, fila, 5, "PROG.\nFINANCIERA\nANUAL", _F_AZUL_OSC, _BLANCO_B, _CENTER)
        _cell(ws, fila, 6, f"PROG.\nACUM.\nENE.", _F_AZUL, _BLANCO_B, _CENTER)
        _cell(ws, fila, 7, f"EJEC.\nACUM.\nENE.", _F_AZUL, _BLANCO_B, _CENTER)
        _cell(ws, fila, 8, "%\nAVANCE", _F_AZUL, _BLANCO_B, _CENTER)
        _cell(ws, fila, 9, f"PROG.\nACUM.\n{mes_nombre}.", _F_AZUL, _BLANCO_B, _CENTER)
        _cell(ws, fila,10, f"EJEC.\nACUM.\n{mes_nombre}.", _F_AZUL, _BLANCO_B, _CENTER)
        _cell(ws, fila,11, "%\nAVANCE", _F_AZUL, _BLANCO_B, _CENTER)
        _cell(ws, fila,12, "PIM", _F_GRIS, _NEGRO_B, _CENTER)
        _cell(ws, fila,13, f"DEVENGADO\n{mes_nombre}", _F_GRIS, _NEGRO_B, _CENTER)
        _cell(ws, fila,14, "%\nEJEC.\nFIN.", _F_GRIS, _NEGRO_B, _CENTER)
        ws.row_dimensions[fila].height = 42
        fila += 1

        for d in acts:
            pct_f = d["pctAvance"]
            fill  = _F_ROJO if (pct_f is not None and pct_f < 50) else (_F_VERDE if (pct_f or 0) >= 80 else None)
            vals  = [None, d["actividadOp"], d["unidadMedida"],
                     d["progAnual"], d["progFinAnual"],
                     d["progMes1"], d["ejecMes1"],
                     f"{d['pctMes1']:.1f}%" if d["pctMes1"] is not None else "—",
                     d["progAcum"], d["ejecAcum"],
                     f"{pct_f:.1f}%" if pct_f is not None else "—",
                     d["pim"], d["devengado"],
                     f"{d['pctFinanciero']:.1f}%" if d["pctFinanciero"] is not None else "—"]
            for c, v in enumerate(vals, 1):
                _cell(ws, fila, c, v, fill, _NORMAL)
            ws.row_dimensions[fila].height = 18
            fila += 1

        fila += 1  # fila vacía entre bloques

    anchos = [10, 50, 8, 12, 14, 12, 12, 9, 12, 12, 9, 14, 14, 9]
    for c, w in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "B2"


def _hoja_tabla02_upe(wb: Workbook, actividades: list, mes: int, mes_nombre: str):
    ws = wb.create_sheet("TABLA 02-UPE")
    fila = 1

    _cell(ws, fila, 1, "TABLERO DE SEGUIMIENTO A LAS UPE", _F_AZUL_OSC, _BLANCO_B)
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=9)
    ws.row_dimensions[fila].height = 20; fila += 2

    for ao in actividades:
        if not ao["upes"]: continue
        # Encabezado del AO
        titulo = f"AO: {ao['codAO']} — {ao['actividadOp']}"
        _cell(ws, fila, 1, titulo, _F_AZUL, _BLANCO_B)
        ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=9)
        ws.row_dimensions[fila].height = 18; fila += 1

        # Cabecera columnas
        for c, txt in enumerate(["N°","UPE",f"PROG. ACUM. ENE-{mes_nombre}",
                                  f"EJEC. ACUM. ENE-{mes_nombre}","% AVANCE",
                                  "PIM","DEVENGADO","U.M."], 1):
            _cell(ws, fila, c, txt, _F_AZUL_CLARO, _NEGRO_B, _CENTER)
        ws.row_dimensions[fila].height = 30; fila += 1

        for n, upe in enumerate(ao["upes"], 1):
            _cell(ws, fila, 1, n, None, _NORMAL, _CENTER)
            _cell(ws, fila, 2, upe["departamento"], None, _NORMAL)
            _cell(ws, fila, 3, upe["progAcum"],  None, _NORMAL, _CENTER)
            _cell(ws, fila, 4, upe["ejecAcum"],  None, _NORMAL, _CENTER)
            _cell(ws, fila, 5,
                  f"{upe['pctAvance']:.1f}%" if upe["pctAvance"] is not None else "—",
                  None, _NORMAL, _CENTER)
            _cell(ws, fila, 6, upe["pim"],       None, _NORMAL, _CENTER)
            _cell(ws, fila, 7, upe["devengado"],  None, _NORMAL, _CENTER)
            _cell(ws, fila, 8, ao["unidadMedida"], None, _NORMAL, _CENTER)
            ws.row_dimensions[fila].height = 15; fila += 1
        fila += 1

    for c, w in enumerate([6, 30, 16, 16, 10, 14, 14, 10], 1):
        ws.column_dimensions[get_column_letter(c)].width = w


def _hoja_pp117_entidad(wb: Workbook, nombre: str, categorias: list,
                         mes: int, mes_nombre: str):
    ws = wb.create_sheet(nombre)
    fila = 1

    # Encabezado general
    hdrs = ["CENTRO DE COSTO", "PP / PRODUCTO / ACTIVIDAD", "U.M."]
    hdrs += [MESES_NOMBRES[i] for i in range(mes)]
    hdrs += [f"PROG. ENE-{mes_nombre}", "TOTAL PROG. ANUAL",
             f"EJEC. ENE-{mes_nombre}", "% AVANCE",
             "PIM", f"DEVENGADO ENE-{mes_nombre}", "% EJEC."]
    for c, h in enumerate(hdrs, 1):
        _cell(ws, fila, c, h, _F_AZUL_OSC, _BLANCO_B, _CENTER)
    ws.row_dimensions[fila].height = 35; fila += 1

    for cat in categorias:
        # Fila Categoría/Producto
        cat_txt = f"{cat['categoriaId']}.{cat['categoria']}" if cat["categoria"] else cat["categoriaId"]
        _cell(ws, fila, 1, cat_txt, _F_AZUL, _BLANCO_B)
        ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=len(hdrs))
        ws.row_dimensions[fila].height = 20; fila += 1

        for ap in cat["aps"]:
            for d in ap["centros"]:
                alias = CC_ALIAS.get(d["centroCosto"], d["centroCosto"])
                ap_txt = f"{d['actPresupId']}.{d['actPresup']}" if d["actPresup"] else d["actPresupId"]
                pct_f  = d["pctAvance"]
                fill   = _F_ROJO if (pct_f is not None and pct_f < 50) else None
                vals   = [alias, ap_txt, d["unidadMedida"]]
                vals  += [d["progMeses"][i] for i in range(mes)]
                vals  += [d["progAcum"], d["progAnual"],
                          d["ejecAcum"],
                          f"{pct_f:.1f}%" if pct_f is not None else "—",
                          d["pim"], d["devengado"],
                          f"{d['pctFinanciero']:.1f}%" if d["pctFinanciero"] is not None else "—"]
                for c, v in enumerate(vals, 1):
                    _cell(ws, fila, c, v, fill, _NORMAL)
                ws.row_dimensions[fila].height = 16; fila += 1

        fila += 1

    for c, w in enumerate([12, 50, 8] + [8]*mes + [12, 14, 12, 9, 14, 14, 9], 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A2"


# ─── GET /descargar/dgnna ────────────────────────────────────────

@router.get("/descargar/dgnna")
def descargar_dgnna(
    mes: int = Query(..., ge=1, le=12),
    anio: int = Query(..., ge=2020),
    db: Session = Depends(get_db),
    _: dict = Depends(require_module_access(MODULO)),
):
    carga = _get_carga_o_404(db, mes, anio)
    mn = MESES_NOMBRES[mes-1]

    filas = db.query(PoiDato).filter(
        PoiDato.cargaId == carga.id,
        PoiDato.centroCosto.in_(CC_DGNNA),
    ).all()
    datos = _agregar_filas(filas, mes)

    # porCC para TABLA 01
    cc_grupos: dict = defaultdict(list)
    for d in datos: cc_grupos[d["centroCosto"]].append(d)
    por_cc = [{"cc": cc, "actividades": acts} for cc, acts in cc_grupos.items()]

    # UPE para TABLA 02
    ao_grupos: dict = defaultdict(lambda: {"codAO":"","actividadOp":"","unidadMedida":"","upes":[]})
    for f in filas:
        key = f.codAO or ""
        g   = ao_grupos[key]
        g["codAO"] = f.codAO or ""; g["actividadOp"] = f.actividadOp or ""
        g["unidadMedida"] = f.unidadMedida or ""
        re_v = [getattr(f,f"fRe{i:02d}") or 0 for i in range(1,13)]
        se_v = [getattr(f,f"fSe{i:02d}") or 0 for i in range(1,13)]
        pr = sum(re_v[:mes]); ej = sum(se_v[:mes])
        g["upes"].append({"departamento": f.departamento or "—",
                           "progAcum": round(pr,2), "ejecAcum": round(ej,2),
                           "pctAvance": _pct(ej,pr),
                           "pim": round(f.consPIM or 0,2),
                           "devengado": round(f.fnSeTotal or 0,2)})

    wb = Workbook(); wb.remove(wb.active)
    _hoja_tabla01(wb, "TABLA 01 DGNNA", por_cc, mes, mn, anio)
    _hoja_tabla02_upe(wb, list(ao_grupos.values()), mes, mn)

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="Reporte-POI-DGNNA-{mn}{anio}.xlsx"'})


# ─── GET /descargar/pp117 ────────────────────────────────────────

@router.get("/descargar/pp117")
def descargar_pp117(
    mes: int = Query(..., ge=1, le=12),
    anio: int = Query(..., ge=2020),
    db: Session = Depends(get_db),
    _: dict = Depends(require_module_access(MODULO)),
):
    carga = _get_carga_o_404(db, mes, anio)
    mn = MESES_NOMBRES[mes-1]

    def _get(cc_f):
        rows = db.query(PoiDato).filter(
            PoiDato.cargaId == carga.id, PoiDato.centroCosto.in_(cc_f)).all()
        datos = _agregar_filas(rows, mes)
        cat_map: dict = defaultdict(lambda: {
            "categoriaId":"","categoria":"",
            "aps": defaultdict(lambda: {"actPresupId":"","actPresup":"","centros":[]})
        })
        for d in datos:
            cat = cat_map[d["categoriaId"] or "S/C"]
            cat["categoriaId"] = d["categoriaId"] or ""; cat["categoria"] = d["categoria"] or ""
            ap  = cat["aps"][d["actPresupId"] or "S/AP"]
            ap["actPresupId"] = d["actPresupId"] or ""; ap["actPresup"] = d["actPresup"] or ""
            ap["centros"].append(d)
        return [{"categoriaId": c["categoriaId"], "categoria": c["categoria"],
                 "aps": [{"actPresupId": a["actPresupId"], "actPresup": a["actPresup"],
                           "centros": a["centros"]} for a in c["aps"].values()]}
                for c in cat_map.values()]

    wb = Workbook(); wb.remove(wb.active)
    _hoja_pp117_entidad(wb, "PP117-SECTOR MIMP", _get(CC_PP117),  mes, mn)
    _hoja_pp117_entidad(wb, "PP117-DGNNA",       _get(CC_DGNNA),  mes, mn)
    _hoja_pp117_entidad(wb, "PP117-INABIF",      _get(CC_INABIF), mes, mn)
    _hoja_pp117_entidad(wb, "PP117-CONADIS",     _get(CC_CONADIS),mes, mn)

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="Reporte-POI-PP0117-{mn}{anio}.xlsx"'})
