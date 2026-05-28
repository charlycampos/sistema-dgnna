"""
MIGRACIÓN DE DATOS: APELA.xlsx → Oracle
========================================
Ejecutar desde la carpeta backend/:
    python migrar_excel_oracle.py

Requisitos:
    - Oracle XE corriendo en localhost:1521
    - .env configurado con DATABASE_URL de Oracle
    - pip install oracledb sqlalchemy openpyxl python-dotenv
    - El archivo APELA.xlsx debe estar en: ../APELA.xlsx
"""

import os
import sys
import uuid
from datetime import datetime
from pathlib import Path

# ── Asegurar que el script corre desde la carpeta backend/ ────────
script_dir = Path(__file__).parent
os.chdir(script_dir)
sys.path.insert(0, str(script_dir))

try:
    import openpyxl
except ImportError:
    print("ERROR: Falta openpyxl. Ejecuta: pip install openpyxl")
    sys.exit(1)

try:
    from database import SessionLocal
    from models import Abogado, ComplejidadJuridica, ExtensionRango, Apelacion
except Exception as e:
    print(f"ERROR al importar módulos del backend: {e}")
    sys.exit(1)

# ── Ruta del Excel ────────────────────────────────────────────────
EXCEL_PATH = script_dir.parent / "APELA.xlsx"
if not EXCEL_PATH.exists():
    print(f"ERROR: No se encontró el archivo Excel en: {EXCEL_PATH}")
    sys.exit(1)

print("=" * 55)
print("  MIGRACIÓN APELA.xlsx → Oracle")
print("=" * 55)
print(f"  Excel:    {EXCEL_PATH}")
print(f"  Base de datos: Oracle (ver .env)")
print("=" * 55)

db = SessionLocal()

# ── Seed de catálogos ─────────────────────────────────────────────
print("\n[1/3] Cargando catálogos base...")

# Abogados
abogados_data = [
    ("Clara Michaud",    True),
    ("Karla Garcia",     True),
    ("Karol Castro",     False),
    ("Sofia Goicochea",  False),
    ("Tula Davila",      False),
]
abogados_map = {}
for nombre, activo in abogados_data:
    a = db.query(Abogado).filter(Abogado.nombre == nombre).first()
    if not a:
        a = Abogado(id=str(uuid.uuid4()), nombre=nombre, activo=activo)
        db.add(a)
        db.flush()
        print(f"      + Abogado creado: {nombre}")
    abogados_map[nombre.upper()] = a.id

# Complejidades jurídicas (nombres exactos del Excel)
comps_data = [
    ("Adopciones",                              1),
    ("UPE (1-2 apelantes)",                    2),
    ("UPE (3 a mas apelantes)",                3),
    ("Procedimiento Administrativo Sancionador", 4),
]
comp_map = {}
for nombre, pts in comps_data:
    c = db.query(ComplejidadJuridica).filter(ComplejidadJuridica.nombre == nombre).first()
    if not c:
        c = ComplejidadJuridica(id=str(uuid.uuid4()), nombre=nombre, puntos=pts)
        db.add(c)
        db.flush()
        print(f"      + Complejidad creada: {nombre} ({pts} pts)")
    comp_map[nombre.upper()] = (c.id, pts)

# Rangos de extensión
rangos_data = [
    ("1 - 500",    1,    500,  1),
    ("501 - 1000", 501,  1000, 2),
    ("1001 - 2000",1001, 2000, 3),
    ("2001 - mas", 2001, None, 4),
]
for desc, minf, maxf, pts in rangos_data:
    if not db.query(ExtensionRango).filter(ExtensionRango.descripcion == desc).first():
        db.add(ExtensionRango(id=str(uuid.uuid4()), descripcion=desc, minFolios=minf, maxFolios=maxf, puntos=pts))
        print(f"      + Rango creado: {desc}")

db.commit()
print("      ✅ Catálogos OK")

# ── Leer Excel ────────────────────────────────────────────────────
print(f"\n[2/3] Leyendo {EXCEL_PATH.name}...")
try:
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb["Registro"]
    print(f"      Filas encontradas: {ws.max_row - 1}")
except Exception as e:
    print(f"      ❌ Error leyendo Excel: {e}")
    db.close()
    sys.exit(1)

# Expedientes ya en BD (para evitar duplicados)
existentes = {r.numeroExpediente for r in db.query(Apelacion.numeroExpediente).all()}
print(f"      Expedientes ya en BD: {len(existentes)}")

# ── Insertar apelaciones ──────────────────────────────────────────
print("\n[3/3] Insertando apelaciones...")

insertados  = 0
skip_vacios = 0
skip_dupes  = 0
errores     = []

for r in range(2, ws.max_row + 1):
    # Número de expediente (col 5) — obligatorio
    exp = ws.cell(r, 5).value
    if not exp:
        skip_vacios += 1
        continue
    exp = str(exp).strip()

    # Saltar duplicados
    if exp in existentes:
        skip_dupes += 1
        continue

    # Abogado (col 15)
    abg_key = str(ws.cell(r, 15).value or "").strip().upper()
    if abg_key not in abogados_map:
        errores.append(f"Fila {r} ({exp}): abogado desconocido [{ws.cell(r,15).value}]")
        continue

    # Complejidad (col 13)
    comp_key = str(ws.cell(r, 13).value or "").strip().upper()
    if comp_key not in comp_map:
        errores.append(f"Fila {r} ({exp}): complejidad desconocida [{ws.cell(r,13).value}]")
        continue

    # Fecha ingreso (col 4) — obligatoria
    fi_raw = ws.cell(r, 4).value
    if not isinstance(fi_raw, datetime):
        errores.append(f"Fila {r} ({exp}): sin fecha de ingreso")
        continue

    # Fechas opcionales
    fim = ws.cell(r, 2).value if isinstance(ws.cell(r, 2).value, datetime) else None
    pv  = ws.cell(r, 3).value if isinstance(ws.cell(r, 3).value, datetime) else None
    fa_raw = ws.cell(r, 16).value
    fa  = fa_raw if isinstance(fa_raw, datetime) else fi_raw

    # Datos principales
    # NOTA: En Oracle '' (cadena vacía) == NULL, así que usamos fallback para campos NOT NULL
    apelante    = str(ws.cell(r, 6).value or "").strip() or "SIN APELANTE"
    nna_car     = str(ws.cell(r, 7).value or "").strip() or None
    procedencia = str(ws.cell(r, 8).value or "").strip() or "SIN PROCEDENCIA"
    documento   = str(ws.cell(r, 9).value or "").strip() or "SIN DOCUMENTO"
    asunto      = str(ws.cell(r, 10).value or "").strip() or "SIN ASUNTO"
    folios      = int(ws.cell(r, 11).value or 0)

    # Puntos (usar valores cachelados del Excel; calcular si falta)
    pe = ws.cell(r, 12).value
    pts_ext  = int(pe) if isinstance(pe, (int, float)) else (
        1 if folios <= 500 else 2 if folios <= 1000 else 3 if folios <= 2000 else 4
    )
    pc = ws.cell(r, 14).value
    pts_comp = int(pc) if isinstance(pc, (int, float)) else comp_map[comp_key][1]
    pt = ws.cell(r, 18).value
    pts_tot  = int(pt) if isinstance(pt, (int, float)) else pts_ext + pts_comp

    # Estado
    estado = str(ws.cell(r, 17).value or "Pendiente").strip()
    if estado not in ("Pendiente", "Resuelto", "Atendido"):
        estado = "Pendiente"

    # Opcionales finales
    num_res = str(ws.cell(r, 19).value or "").strip() or None
    doc_aten = str(ws.cell(r, 20).value or "").strip() or None
    cargos   = str(ws.cell(r, 21).value or "").strip() or None

    try:
        ap = Apelacion(
            id                = str(uuid.uuid4()),
            numeroExpediente  = exp,
            fechaIngreso      = fi_raw,
            fechaIngresoMIMP  = fim,
            plazoVencimiento  = pv,
            apelante          = apelante,
            nnaCar            = nna_car,
            procedencia       = procedencia,
            documento         = documento,
            asunto            = asunto,
            folios            = folios,
            puntosExtension   = pts_ext,
            complejidadId     = comp_map[comp_key][0],
            puntosComplejidad = pts_comp,
            puntosTotal       = pts_tot,
            abogadoId         = abogados_map[abg_key],
            fechaAsignacion   = fa,
            estado            = estado,
            numeroResolucion  = num_res,
            documentoAtencion = doc_aten,
            cargos            = cargos,
        )
        db.add(ap)
        db.flush()          # detecta error Oracle antes del commit
        existentes.add(exp)
        insertados += 1

        # Commit cada 50 registros
        if insertados % 50 == 0:
            db.commit()
            print(f"      ... {insertados} registros insertados")

    except Exception as e:
        db.rollback()       # limpiar sesión para poder continuar
        errores.append(f"Fila {r} ({exp}): {e}")

db.commit()
db.close()

# ── Resumen ───────────────────────────────────────────────────────
print("\n" + "=" * 55)
print("  RESULTADO FINAL")
print("=" * 55)
print(f"  ✅ Insertados:              {insertados}")
print(f"     Sin expediente (omit.): {skip_vacios}")
print(f"     Duplicados (omit.):     {skip_dupes}")
if errores:
    print(f"\n  ⚠️  Errores ({len(errores)}):")
    for e in errores[:20]:
        print(f"     - {e}")
else:
    print(f"  ✅ Sin errores")
print("=" * 55)
print("\n  Migración completada. Reinicia el servidor backend.")
print("  Luego recarga el módulo de apelaciones en el navegador.")
